"""
Summer Strength Monitor v4 — India Consumer Equities
======================================================
Jan-Jun full pre-monsoon season tracker with:
- Summer Onset Tracker (first 30°C/35°C crossover per city, YoY)
- Heating Rate (weekly temperature climb vs normal)
- Dual threshold cumulative chart (≥30°C + ≥35°C city-days)
- Monthly heatmap (cities × months, deviation from normal)
- 5yr baseline deviation, YoY comparison, SSI

Usage:  python summer_monitor.py [--validate | -v]
Output: summer_strength_monitor.html
Needs:  pip install requests
"""
import requests, json, os, sys, time
from datetime import date, datetime
from pathlib import Path

# ═══════════════════════════════════════════════════════════════
# CONFIGURATION
# ═══════════════════════════════════════════════════════════════
CITIES = [
    # ── North India ──
    {"name":"Delhi",         "lat":28.58,"lon":77.21,"region":"North"},
    {"name":"Jaipur",        "lat":26.82,"lon":75.80,"region":"North"},
    {"name":"Lucknow",       "lat":26.85,"lon":80.95,"region":"North"},
    {"name":"Kanpur",        "lat":26.45,"lon":80.35,"region":"North"},
    {"name":"Chandigarh",    "lat":30.73,"lon":76.78,"region":"North"},
    {"name":"Varanasi",      "lat":25.32,"lon":83.01,"region":"North"},    # UP heat belt — VBL heartland
    {"name":"Prayagraj",     "lat":25.43,"lon":81.85,"region":"North"},    # Gangetic plain extreme heat
    {"name":"Gorakhpur",     "lat":26.76,"lon":83.37,"region":"North"},    # Eastern UP, routinely 45°C+
    {"name":"Jodhpur",       "lat":26.29,"lon":73.02,"region":"North"},    # Thar desert edge, peak heat
    {"name":"Bikaner",       "lat":28.02,"lon":73.31,"region":"North"},    # One of India's hottest cities
    # ── West India ──
    {"name":"Mumbai",        "lat":19.08,"lon":72.88,"region":"West","source":"vc"},  # Visual Crossing (METAR Santa Cruz) — Open-Meteo grid can't resolve peninsula
    {"name":"Ahmedabad",     "lat":23.07,"lon":72.63,"region":"West"},
    {"name":"Pune",          "lat":18.52,"lon":73.86,"region":"West"},
    {"name":"Vadodara",      "lat":22.31,"lon":73.19,"region":"West"},
    {"name":"Surat",         "lat":21.17,"lon":72.83,"region":"West"},
    {"name":"Rajkot",        "lat":22.30,"lon":70.80,"region":"West"},     # Saurashtra — heatwave alert zone
    # ── Central India ──
    {"name":"Nagpur",        "lat":21.10,"lon":79.05,"region":"Central"},
    {"name":"Indore",        "lat":22.72,"lon":75.86,"region":"Central"},
    {"name":"Bhopal",        "lat":23.26,"lon":77.41,"region":"Central"},
    {"name":"Raipur",        "lat":21.25,"lon":81.63,"region":"Central"},  # Chhattisgarh, extreme heat
    # ── South India ──
    {"name":"Bangalore",     "lat":12.97,"lon":77.59,"region":"South"},
    {"name":"Hyderabad",     "lat":17.45,"lon":78.47,"region":"South"},
    {"name":"Chennai",       "lat":13.00,"lon":80.18,"region":"South"},
    {"name":"Coimbatore",    "lat":11.02,"lon":76.96,"region":"South"},
    {"name":"Madurai",       "lat":9.92, "lon":78.12,"region":"South"},    # Interior TN, much hotter than Chennai
    {"name":"Vijayawada",    "lat":16.51,"lon":80.65,"region":"South"},    # AP interior, extreme heat zone
    # ── East India ──
    {"name":"Kolkata",       "lat":22.65,"lon":88.45,"region":"East"},
    {"name":"Patna",         "lat":25.61,"lon":85.10,"region":"East"},
    {"name":"Visakhapatnam", "lat":17.69,"lon":83.22,"region":"East"},
    {"name":"Bhubaneswar",   "lat":20.30,"lon":85.83,"region":"East"},    # Odisha — regular heatwave zone
]


OUTPUT_FILE = "summer_strength_monitor.html"
VC_KEY = None  # Set via VC_KEY env var or --vc-key arg

# ═══════════════════════════════════════════════════════════════
# DATA FETCHING
# ═══════════════════════════════════════════════════════════════
def get_date_ranges():
    today = date.today()
    year = today.year
    season_start = date(year, 1, 1)
    season_end = date(year, 6, 30)
    if today < season_start:
        year -= 1; season_start = date(year,1,1); season_end = date(year,6,30); today = season_end
    end_current = min(today, season_end)
    ly = year - 1
    return {
        "start_current": season_start.isoformat(),
        "end_current": end_current.isoformat(),
        "start_last": date(ly,1,1).isoformat(),
        "end_last_compare": date(ly, end_current.month, end_current.day).isoformat(),
        "end_last_full": date(ly,6,30).isoformat(),
        "year": year, "last_year": ly,
        "days_tracked": (end_current - season_start).days + 1,
        "baseline_start": date(year-6,1,1).isoformat(),
        "baseline_end": date(year-2,6,30).isoformat(),
        "baseline_label": f"{year-6}-{year-2}",
    }

def fetch_om(lat, lon, start, end, archive=False):
    base = "https://archive-api.open-meteo.com/v1/archive" if archive else "https://api.open-meteo.com/v1/forecast"
    r = requests.get(base, params={"latitude":lat,"longitude":lon,"start_date":start,"end_date":end,
        "daily":"temperature_2m_max,temperature_2m_min,precipitation_sum","timezone":"Asia/Kolkata"}, timeout=30)
    r.raise_for_status(); d = r.json()
    if "error" in d: raise Exception(d.get("reason",""))
    return d.get("daily",{}), {"lat":d.get("latitude"),"lon":d.get("longitude"),"elevation":d.get("elevation")}

def fetch_vc(location, start, end):
    """Fetch from Visual Crossing (METAR station data). Used for Mumbai."""
    if not VC_KEY:
        raise Exception("No Visual Crossing API key. Use --vc-key KEY or set VC_KEY env var.")
    url = f"https://weather.visualcrossing.com/VisualCrossingWebServices/rest/services/timeline/{location}/{start}/{end}"
    r = requests.get(url, params={
        "unitGroup": "metric", "key": VC_KEY,
        "include": "days", "elements": "datetime,tempmax,tempmin,precip",
    }, timeout=30)
    r.raise_for_status()
    d = r.json()
    days = d.get("days", [])
    # Convert to Open-Meteo-compatible format
    daily = {
        "time": [day["datetime"] for day in days],
        "temperature_2m_max": [day.get("tempmax") for day in days],
        "temperature_2m_min": [day.get("tempmin") for day in days],
        "precipitation_sum": [day.get("precip", 0) or 0 for day in days],
    }
    resolved = {"lat": d.get("latitude"), "lon": d.get("longitude"), "elevation": None}
    return daily, resolved

# ─── Visual Crossing cache (avoids burning daily quota on repeat runs) ───────
def _vc_cache_path(dates):
    return Path(__file__).parent / f"mumbai_vc_cache_{dates['year']}_{dates['end_current']}.json"

def _load_vc_cache(dates):
    p = _vc_cache_path(dates)
    if not p.exists(): return None
    age_hours = (time.time() - p.stat().st_mtime) / 3600
    if age_hours > 23: return None          # stale — re-fetch after 23 h
    try:
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None

def _save_vc_cache(dates, payload):
    p = _vc_cache_path(dates)
    try:
        with open(p, "w", encoding="utf-8") as f:
            json.dump(payload, f)
        # Clean up cache files from previous days/seasons
        for old in p.parent.glob("mumbai_vc_cache_*.json"):
            if old != p:
                try: old.unlink()
                except Exception: pass
    except Exception as e:
        print(f"  [cache write failed: {e}]")

def fetch_city(city, dates):
    lat, lon = city["lat"], city["lon"]
    is_vc = city.get("source") == "vc"
    use_arch = (date.today() - date.fromisoformat(dates["start_current"])).days > 90
    resolved = None

    if is_vc and VC_KEY:
        # Visual Crossing for Mumbai — uses "Mumbai,India" location string
        loc = f"{city['name']},India"

        # ── Try cache first (saves VC daily quota on repeat runs) ──
        cached = _load_vc_cache(dates)
        if cached:
            print(f"[CACHE] ", end="")
            return cached

        # ── Live fetch ──
        try: curr, resolved = fetch_vc(loc, dates["start_current"], dates["end_current"])
        except Exception as e:
            print(f"[VC error: {e}] ", end=""); curr = None
        try: lc, _ = fetch_vc(loc, dates["start_last"], dates["end_last_compare"])
        except: lc = None
        try: lf, _ = fetch_vc(loc, dates["start_last"], dates["end_last_full"])
        except: lf = None
        # VC baseline: fetch year-by-year (VC caps single requests at ~365 days)
        # Use 3 years: yr-4, yr-3, yr-2 (each Jan 1 - Jun 30 = ~181 records)
        yr = dates["year"]
        bl = None
        try:
            bl_times, bl_tmax, bl_tmin, bl_rain = [], [], [], []
            for by in range(yr-4, yr-1):  # e.g. 2022, 2023, 2024
                bdata, _ = fetch_vc(loc, f"{by}-01-01", f"{by}-06-30")
                if bdata:
                    bl_times.extend(bdata.get("time",[]))
                    bl_tmax.extend(bdata.get("temperature_2m_max",[]))
                    bl_tmin.extend(bdata.get("temperature_2m_min",[]))
                    bl_rain.extend(bdata.get("precipitation_sum",[]))
                time.sleep(0.5)
            if bl_times:
                bl = {"time":bl_times,"temperature_2m_max":bl_tmax,"temperature_2m_min":bl_tmin,"precipitation_sum":bl_rain}
        except: bl = None
    elif is_vc and not VC_KEY:
        # No VC key — skip Mumbai entirely (Open-Meteo can't resolve it)
        return {"city":city["name"],"region":city["region"],"current":None,"lastYearCompare":None,
                "lastYearFull":None,"baseline":None,"requested_lat":lat,"requested_lon":lon,
                "resolved":None,"source":"skipped"}
    else:
        # Open-Meteo for all other cities
        try: curr, resolved = fetch_om(lat, lon, dates["start_current"], dates["end_current"], use_arch)
        except:
            try: curr, resolved = fetch_om(lat, lon, dates["start_current"], dates["end_current"], not use_arch)
            except: curr = None
        try: lc, _ = fetch_om(lat, lon, dates["start_last"], dates["end_last_compare"], True)
        except: lc = None
        try: lf, _ = fetch_om(lat, lon, dates["start_last"], dates["end_last_full"], True)
        except: lf = None
        try: bl, _ = fetch_om(lat, lon, dates["baseline_start"], dates["baseline_end"], True)
        except: bl = None

    result = {"city":city["name"],"region":city["region"],"current":curr,"lastYearCompare":lc,
            "lastYearFull":lf,"baseline":bl,"requested_lat":lat,"requested_lon":lon,"resolved":resolved,
            "source": "vc" if (is_vc and VC_KEY) else "om"}
    if is_vc and VC_KEY and curr is not None:
        _save_vc_cache(dates, result)
    return result

def fetch_all(dates):
    results = {}
    for i, c in enumerate(CITIES):
        is_vc = c.get("source") == "vc" and VC_KEY
        skip_vc = c.get("source") == "vc" and not VC_KEY
        src_tag = "VC" if is_vc else ("SKIP" if skip_vc else "OM")
        print(f"  [{i+1}/{len(CITIES)}] {c['name']} [{src_tag}]...", end=" ", flush=True)
        if skip_vc:
            print("SKIPPED (no VC key)")
            continue
        try:
            d = fetch_city(c, dates)
            n = len(d["current"]["temperature_2m_max"]) if d["current"] else 0
            r = d.get("resolved")
            if r and not is_vc:
                drift = ((r["lat"]-c["lat"])**2+(r["lon"]-c["lon"])**2)**0.5 * 111
                info = f"resolved: {r['lat']:.2f},{r['lon']:.2f} elev:{r['elevation']}m"
                if drift > 5: info += f" ⚠ {drift:.0f}km drift"
                print(f"OK ({n}d) [{info}]")
            elif is_vc:
                print(f"OK ({n}d) [Visual Crossing METAR — Santa Cruz station]")
            else: print(f"OK ({n}d)")
            results[c["name"]] = d
        except Exception as e:
            print(f"FAIL: {e}")
            results[c["name"]] = {"city":c["name"],"region":c["region"],"current":None,
                "lastYearCompare":None,"lastYearFull":None,"baseline":None,
                "requested_lat":c["lat"],"requested_lon":c["lon"],"resolved":None,"source":"failed"}
        time.sleep(0.3 if not is_vc else 0.5)  # slightly slower for VC to respect rate limits
    return results

# ═══════════════════════════════════════════════════════════════
# ANALYTICS
# ═══════════════════════════════════════════════════════════════
def sa(arr):
    v=[x for x in (arr or []) if x is not None]; return sum(v)/len(v) if v else 0
def ss(arr): return sum(x for x in (arr or []) if x is not None)

def first_cross(temps, dates_list, threshold):
    """Find first date where temp crosses threshold."""
    for i, t in enumerate(temps):
        if t is not None and t >= threshold:
            return dates_list[i] if i < len(dates_list) else f"day-{i}"
    return None

def compute_stats(data):
    if not data["current"] or not data["lastYearCompare"]: return None
    curr, prev = data["current"], data["lastYearCompare"]
    tmax_c, tmax_p = curr["temperature_2m_max"], prev["temperature_2m_max"]
    dates_c = curr.get("time",[])
    dates_p = prev.get("time",[])

    avg_c, avg_p = sa(tmax_c), sa(tmax_p)
    peak_c = max((t for t in tmax_c if t is not None), default=0)
    peak_p = max((t for t in tmax_p if t is not None), default=0)
    h35c = sum(1 for t in tmax_c if t and t>=35)
    h35p = sum(1 for t in tmax_p if t and t>=35)
    h40c = sum(1 for t in tmax_c if t and t>=40)
    h40p = sum(1 for t in tmax_p if t and t>=40)
    h30c = sum(1 for t in tmax_c if t and t>=30)
    h30p = sum(1 for t in tmax_p if t and t>=30)
    rain_c = ss(curr.get("precipitation_sum",[]))
    rain_p = ss(prev.get("precipitation_sum",[]))
    delta = avg_c - avg_p

    # SSI (computed on Mar-Jun subset only for meaningful comparison)
    mar_idx_c = next((i for i,d in enumerate(dates_c) if d and d[5:7]>='03'), len(dates_c))
    mar_tmax_c = tmax_c[mar_idx_c:]
    mar_idx_p = next((i for i,d in enumerate(dates_p) if d and d[5:7]>='03'), len(dates_p))
    mar_tmax_p = tmax_p[mar_idx_p:]
    mar_avg_c = sa(mar_tmax_c) if mar_tmax_c else avg_c
    mar_avg_p = sa(mar_tmax_p) if mar_tmax_p else avg_p
    mar_h35c = sum(1 for t in mar_tmax_c if t and t>=35)
    mar_h35p = sum(1 for t in mar_tmax_p if t and t>=35)
    mar_rain_c = ss([curr.get("precipitation_sum",[])[i] for i in range(mar_idx_c, len(tmax_c)) if i < len(curr.get("precipitation_sum",[]))])
    mar_rain_p = ss([prev.get("precipitation_sum",[])[i] for i in range(mar_idx_p, len(tmax_p)) if i < len(prev.get("precipitation_sum",[]))])

    # SSI Components — all capped to prevent small-number distortion
    # 1. Temperature ratio (50%) — most stable, rarely distorts
    ts = mar_avg_c / max(mar_avg_p, 1)
    # 2. Hot days score (30%) — cap ratio between 0.5x and 2.0x
    #    Raw ratio blows up when base is small (e.g. 2→7 = 3.5x for Mumbai)
    #    Capping at 2.0x ensures no single city dominates
    if mar_h35p >= 5:
        # Enough base days — ratio is meaningful
        hs = max(0.5, min(2.0, mar_h35c / mar_h35p))
    elif mar_h35c > mar_h35p:
        # Small base but more hot days this year — mild positive signal
        hs = min(1.3, 1.0 + (mar_h35c - mar_h35p) * 0.05)
    elif mar_h35c < mar_h35p:
        # Fewer hot days — mild negative signal
        hs = max(0.7, 1.0 - (mar_h35p - mar_h35c) * 0.05)
    else:
        hs = 1.0
    # 3. Rainfall score (20%) — less rain = drier = more beverages, capped at 1.5x
    rs = (mar_rain_p / max(mar_rain_c, 0.1)) if mar_rain_p > 0 else (1.1 if mar_rain_c < 1 else 0.9)
    ssi = ts*0.5 + max(0.5, min(2.0, hs))*0.3 + min(rs,1.5)*0.2
    signal = "STRONGER" if ssi>1.05 else ("WEAKER" if ssi<0.95 else "IN-LINE")

    # Mar-onwards additional metrics for summary cards
    mar_h40c = sum(1 for t in mar_tmax_c if t and t>=40)
    mar_h40p = sum(1 for t in mar_tmax_p if t and t>=40)
    mar_h30c = sum(1 for t in mar_tmax_c if t and t>=30)
    mar_h30p = sum(1 for t in mar_tmax_p if t and t>=30)
    mar_peak_c = max((t for t in mar_tmax_c if t is not None), default=0) if mar_tmax_c else 0
    mar_peak_p = max((t for t in mar_tmax_p if t is not None), default=0) if mar_tmax_p else 0
    mar_delta = mar_avg_c - mar_avg_p
    mar_days = len(mar_tmax_c)

    pf = data.get("lastYearFull") or {}
    pf_tmax = pf.get("temperature_2m_max",[])
    pf_dates = pf.get("time",[])

    # ── Summer Onset ──
    onset30_c = first_cross(tmax_c, dates_c, 30)
    onset35_c = first_cross(tmax_c, dates_c, 35)
    onset30_p = first_cross(tmax_p, dates_p, 30)
    onset35_p = first_cross(tmax_p, dates_p, 35)

    def day_of_year(ds):
        if not ds: return None
        try: return (date.fromisoformat(ds) - date(int(ds[:4]),1,1)).days
        except: return None

    onset30_doy_c = day_of_year(onset30_c)
    onset30_doy_p = day_of_year(onset30_p)
    onset35_doy_c = day_of_year(onset35_c)
    onset35_doy_p = day_of_year(onset35_p)
    onset30_delta = (onset30_doy_c - onset30_doy_p) if onset30_doy_c is not None and onset30_doy_p is not None else None
    onset35_delta = (onset35_doy_c - onset35_doy_p) if onset35_doy_c is not None and onset35_doy_p is not None else None

    # ── Heating Rate (weekly avg max, for each week) ──
    def weekly_avgs(temps):
        weeks = []
        for w in range(0, len(temps), 7):
            chunk = [t for t in temps[w:w+7] if t is not None]
            weeks.append(round(sum(chunk)/len(chunk),1) if chunk else None)
        return weeks
    weekly_c = weekly_avgs(tmax_c)
    weekly_p = weekly_avgs(tmax_p)

    # ── 5yr Baseline Normal ──
    bl = data.get("baseline")
    normal_by_day = {}
    if bl and bl.get("time") and bl.get("temperature_2m_max"):
        for ds, t in zip(bl["time"], bl["temperature_2m_max"]):
            if t is None: continue
            mm, dd = int(ds[5:7]), int(ds[8:10])
            if mm > 6: continue
            doy = (date(2000,mm,dd) - date(2000,1,1)).days
            if doy not in normal_by_day: normal_by_day[doy] = []
            normal_by_day[doy].append(t)

    n_days = len(tmax_c)
    daily_normals = []
    for i in range(182):  # Jan 1 - Jun 30 = 181 days
        if i in normal_by_day and normal_by_day[i]:
            daily_normals.append(round(sum(normal_by_day[i])/len(normal_by_day[i]),1))
        else: daily_normals.append(None)

    # Normal for comparable period (full Jan-today)
    comp_normals = daily_normals[:n_days]
    avg_normal = sa(comp_normals)
    dev = round(avg_c - avg_normal, 1) if avg_normal else None

    # Normal for Mar onwards only (for summary cards)
    mar_doy_start = 59  # Mar 1 = day 59 (Jan=31 + Feb=28)
    mar_normals = daily_normals[mar_doy_start:n_days] if n_days > mar_doy_start else []
    mar_avg_normal = sa(mar_normals) if mar_normals else None
    mar_dev = round(mar_avg_c - mar_avg_normal, 1) if mar_avg_normal else None
    nsig = "WELL ABOVE" if dev and dev>=3 else ("ABOVE" if dev and dev>=1.5 else ("BELOW" if dev and dev<=-1.5 else "NEAR NORMAL"))

    # ── Weekly normals for heating rate comparison ──
    weekly_norm = weekly_avgs(daily_normals[:n_days])

    # ── Cumulative hot days (running totals) ──
    def cum_count(temps, threshold):
        out, r = [], 0
        for t in temps:
            if t is not None and t >= threshold: r += 1
            out.append(r)
        return out
    cum30_c = cum_count(tmax_c, 30)
    cum35_c = cum_count(tmax_c, 35)
    cum30_pf = cum_count(pf_tmax, 30)
    cum35_pf = cum_count(pf_tmax, 35)

    # ── Monthly averages (for heatmap) ──
    def monthly_avg(temps, dates_list):
        months = {}
        for d, t in zip(dates_list, temps):
            if t is None or not d: continue
            m = int(d[5:7])
            if m not in months: months[m] = []
            months[m].append(t)
        return {m: round(sum(v)/len(v),1) for m,v in months.items()}

    monthly_c = monthly_avg(tmax_c, dates_c)
    monthly_p = monthly_avg(tmax_p, dates_p)

    # Monthly normals
    monthly_norm = {}
    for m in range(1,7):
        doy_start = (date(2000,m,1)-date(2000,1,1)).days
        doy_end = (date(2000,m+1,1)-date(2000,1,1)).days if m<6 else 181
        vals = [daily_normals[d] for d in range(doy_start, min(doy_end, len(daily_normals))) if d<len(daily_normals) and daily_normals[d] is not None]
        monthly_norm[m] = round(sum(vals)/len(vals),1) if vals else None

    monthly_dev_c = {m: round(monthly_c.get(m,0)-(monthly_norm.get(m) or 0),1) if monthly_norm.get(m) and m in monthly_c else None for m in range(1,7)}

    return {
        "city": data["city"], "region": data["region"],
        "avg_max_curr": round(avg_c,1), "avg_max_prev": round(avg_p,1),
        "peak_curr": round(peak_c,1), "peak_prev": round(peak_p,1),
        "hot35_curr": h35c, "hot35_prev": h35p,
        "hot40_curr": h40c, "hot40_prev": h40p,
        "hot30_curr": h30c, "hot30_prev": h30p,
        "rain_curr": round(rain_c,1), "rain_prev": round(rain_p,1),
        "delta": round(delta,1), "ssi": round(ssi,3), "signal": signal,
        "days_counted": n_days,
        # Mar-onwards metrics (for summary cards)
        "mar_avg_curr": round(mar_avg_c,1), "mar_avg_prev": round(mar_avg_p,1),
        "mar_delta": round(mar_delta,1), "mar_days": mar_days,
        "mar_h35c": mar_h35c, "mar_h35p": mar_h35p,
        "mar_h40c": mar_h40c, "mar_h40p": mar_h40p,
        "mar_h30c": mar_h30c, "mar_h30p": mar_h30p,
        "mar_peak_curr": round(mar_peak_c,1), "mar_peak_prev": round(mar_peak_p,1),
        "mar_avg_normal": round(mar_avg_normal,1) if mar_avg_normal else None,
        "mar_dev": mar_dev,
        # Daily arrays
        "daily_max_curr": tmax_c, "daily_max_prev_full": pf_tmax,
        "daily_rain_curr": curr.get("precipitation_sum",[]),
        "daily_rain_prev_full": pf.get("precipitation_sum",[]),
        "daily_normals_full": daily_normals,
        # Onset
        "onset30_curr": onset30_c, "onset35_curr": onset35_c,
        "onset30_prev": onset30_p, "onset35_prev": onset35_p,
        "onset30_delta": onset30_delta, "onset35_delta": onset35_delta,
        # Heating rate
        "weekly_curr": weekly_c, "weekly_prev": weekly_p, "weekly_norm": weekly_norm,
        # Cumulative
        "cum30_curr": cum30_c, "cum35_curr": cum35_c,
        "cum30_prev_full": cum30_pf, "cum35_prev_full": cum35_pf,
        # Normal (full Jan-today)
        "avg_normal": round(avg_normal,1) if avg_normal else None,
        "dev_from_normal": dev, "normal_signal": nsig,
        # Monthly heatmap
        "monthly_curr": monthly_c, "monthly_prev": monthly_p,
        "monthly_norm": monthly_norm, "monthly_dev": monthly_dev_c,
    }

# ═══════════════════════════════════════════════════════════════
# HTML GENERATION
# ═══════════════════════════════════════════════════════════════
def generate_html(stats, dates):
    # Full Jan-today aggregates (for charts, onset, heatmap)
    pac = round(sa([s["avg_max_curr"] for s in stats]),1)
    pap = round(sa([s["avg_max_prev"] for s in stats]),1)
    pd = round(pac-pap,1)
    t35c,t35p = sum(s["hot35_curr"] for s in stats), sum(s["hot35_prev"] for s in stats)
    t40c,t40p = sum(s["hot40_curr"] for s in stats), sum(s["hot40_prev"] for s in stats)
    t30c,t30p = sum(s["hot30_curr"] for s in stats), sum(s["hot30_prev"] for s in stats)
    arc = round(sa([s["rain_curr"] for s in stats]),1)
    arp = round(sa([s["rain_prev"] for s in stats]),1)
    rd = round(arc-arp,1)
    pssi = round(sa([s["ssi"] for s in stats]),3)
    ns = sum(1 for s in stats if s["signal"]=="STRONGER")
    nw = sum(1 for s in stats if s["signal"]=="WEAKER")
    ni = len(stats)-ns-nw
    days = stats[0]["days_counted"]

    # Mar-onwards aggregates (for summary cards — reflects summer, not winter)
    mpac = round(sa([s["mar_avg_curr"] for s in stats]),1)
    mpap = round(sa([s["mar_avg_prev"] for s in stats]),1)
    mpd = round(mpac-mpap,1)
    mt35c = sum(s["mar_h35c"] for s in stats)
    mt35p = sum(s["mar_h35p"] for s in stats)
    mt40c = sum(s["mar_h40c"] for s in stats)
    mt40p = sum(s["mar_h40p"] for s in stats)
    mt30c = sum(s["mar_h30c"] for s in stats)
    mt30p = sum(s["mar_h30p"] for s in stats)
    mdays = stats[0]["mar_days"] if stats else 0
    mpan_norm = round(sa([s["mar_avg_normal"] for s in stats if s["mar_avg_normal"]]),1)
    mpan_dev = round(mpac - mpan_norm, 1) if mpan_norm else 0
    mn_above = sum(1 for s in stats if s["mar_dev"] is not None and s["mar_dev"] >= 1.5)

    if pssi>1.05: ve,vt,vc = "🔥","Materially Stronger Summer — Bullish for VBL, Dabur, Havells","#F87171"
    elif pssi<0.95: ve,vt,vc = "❄️","Weaker Summer — Cautious on Summer Plays","#60A5FA"
    else: ve,vt,vc = "➡️","In-Line Summer — No Clear Alpha Signal","#A8A29E"

    ndc = max(len(s["daily_max_curr"]) for s in stats)
    ndp = max(len(s["daily_max_prev_full"]) for s in stats)

    def davg(key,n):
        return [round(sa([s[key][i] for s in stats if i<len(s[key])]),1) if any(i<len(s[key]) and s[key][i] is not None for s in stats) else None for i in range(n)]
    def dcnt(key,n,thr):
        return [sum(1 for s in stats if i<len(s[key]) and s[key][i] is not None and s[key][i]>=thr) for i in range(n)]
    def dcum(key,n):
        return [sum(s[key][i] for s in stats if i<len(s[key])) for i in range(n)]

    ca_c, ca_p = davg("daily_max_curr",ndc), davg("daily_max_prev_full",ndp)
    ch_c, ch_p = dcnt("daily_max_curr",ndc,35), dcnt("daily_max_prev_full",ndp,35)
    cr_c, cr_p = davg("daily_rain_curr",ndc), davg("daily_rain_prev_full",ndp)

    # Normal line
    ndn = max(len(s["daily_normals_full"]) for s in stats)
    cn = davg("daily_normals_full", ndn)

    # Cumulative dual threshold
    cc30c, cc35c = dcum("cum30_curr",ndc), dcum("cum35_curr",ndc)
    cc30p, cc35p = dcum("cum30_prev_full",ndp), dcum("cum35_prev_full",ndp)

    # Heating rate (pan-India weekly avgs)
    nw_c = max(len(s["weekly_curr"]) for s in stats)
    nw_p = max(len(s["weekly_prev"]) for s in stats)
    nw_n = max(len(s["weekly_norm"]) for s in stats)
    def wavg(key,n):
        return [round(sa([s[key][i] for s in stats if i<len(s[key])]),1) if any(i<len(s[key]) and s[key][i] is not None for s in stats) else None for i in range(n)]
    hw_c, hw_p, hw_n = wavg("weekly_curr",nw_c), wavg("weekly_prev",nw_p), wavg("weekly_norm",nw_n)

    # Normal deviation aggregates
    pan_norm = round(sa([s["avg_normal"] for s in stats if s["avg_normal"]]),1)
    pan_dev = round(pac - pan_norm,1) if pan_norm else 0
    n_above = sum(1 for s in stats if s["normal_signal"] in ("WELL ABOVE","ABOVE"))

    # Onset summary
    onset30_deltas = [s["onset30_delta"] for s in stats if s["onset30_delta"] is not None]
    onset35_deltas = [s["onset35_delta"] for s in stats if s["onset35_delta"] is not None]
    avg_onset30_delta = round(sa(onset30_deltas),0) if onset30_deltas else None
    avg_onset35_delta = round(sa(onset35_deltas),0) if onset35_deltas else None
    earlier30 = sum(1 for d in onset30_deltas if d < 0)
    earlier35 = sum(1 for d in onset35_deltas if d < 0)

    sorted_stats = sorted(stats, key=lambda s: s["delta"], reverse=True)
    cj = json.dumps([{
        "city":s["city"],"region":s["region"],
        "avgMaxCurr":s["avg_max_curr"],"avgMaxPrev":s["avg_max_prev"],
        "peakCurr":s["peak_curr"],"peakPrev":s["peak_prev"],
        "delta":s["delta"],"hot35Curr":s["hot35_curr"],"hot35Prev":s["hot35_prev"],
        "hot40Curr":s["hot40_curr"],"hot30Curr":s["hot30_curr"],"hot30Prev":s["hot30_prev"],
        "rainCurr":s["rain_curr"],"rainPrev":s["rain_prev"],
        "ssi":s["ssi"],"signal":s["signal"],
        "avgNormal":s["avg_normal"],"devFromNormal":s["dev_from_normal"],"normalSignal":s["normal_signal"],
        "marAvgCurr":s["mar_avg_curr"],"marAvgPrev":s["mar_avg_prev"],"marDelta":s["mar_delta"],
        "marDev":s["mar_dev"],"marAvgNormal":s["mar_avg_normal"],
        "onset30Curr":s["onset30_curr"],"onset35Curr":s["onset35_curr"],
        "onset30Prev":s["onset30_prev"],"onset35Prev":s["onset35_prev"],
        "onset30Delta":s["onset30_delta"],"onset35Delta":s["onset35_delta"],
        # Aliases for embedded template
        "onset30c":s["onset30_curr"],"onset30p":s["onset30_prev"],"onset30diff":s["onset30_delta"],
        "onset35c":s["onset35_curr"],"onset35p":s["onset35_prev"],"onset35diff":s["onset35_delta"],
        "monthlyDev":{str(m):v for m,v in s["monthly_dev"].items() if v is not None},
        "monthlyCurr":{str(m):v for m,v in s["monthly_curr"].items()},
        "monthlyNorm":{str(m):v for m,v in s["monthly_norm"].items() if v is not None},
    } for s in sorted_stats])

    yr, ly = dates["year"], dates["last_year"]
    gen = datetime.now().strftime("%Y-%m-%d %H:%M IST")

    db = json.dumps({
        "year":yr,"lastYear":ly,"startCurrent":dates["start_current"],"endCurrent":dates["end_current"],
        "daysTracked":days,"numCities":len(stats),
        # Full Jan-today (for charts)
        "panAvgCurr":pac,"panAvgPrev":pap,"panDelta":pd,
        "hot35Curr":t35c,"hot35Prev":t35p,"hot40Curr":t40c,"hot40Prev":t40p,
        "hot30Curr":t30c,"hot30Prev":t30p,
        "avgRainCurr":arc,"avgRainPrev":arp,"rainDelta":rd,
        # Mar-onwards (for summary cards)
        "marAvgCurr":mpac,"marAvgPrev":mpap,"marDelta":mpd,"marDays":mdays,
        "marH35Curr":mt35c,"marH35Prev":mt35p,"marH40Curr":mt40c,"marH40Prev":mt40p,
        "marH30Curr":mt30c,"marH30Prev":mt30p,
        "marAvgNormal":mpan_norm,"marDev":mpan_dev,"marAboveNormal":mn_above,
        "panSSI":pssi,"strongerCount":ns,"weakerCount":nw,"inlineCount":ni,
        "verdictEmoji":ve,"verdictText":vt,"verdictColor":vc,
        "ssiSignal":"STRONGER" if pssi>1.05 else ("WEAKER" if pssi<0.95 else "IN-LINE"),
        "ssiBias":"Bullish" if pssi>1.05 else ("Cautious" if pssi<0.95 else "Neutral"),
        "generatedAt":gen,
        "chartAvgCurr":ca_c,"chartAvgPrev":ca_p,"chartNormal":cn,
        "chartHotCurr":ch_c,"chartHotPrev":ch_p,
        "chartHot35Curr":ch_c,"chartHot35Prev":ch_p,
        "chartRainCurr":cr_c,"chartRainPrev":cr_p,
        "chartCum30Curr":cc30c,"chartCum35Curr":cc35c,
        "chartCum30Prev":cc30p,"chartCum35Prev":cc35p,
        "heatRateCurr":hw_c,"heatRatePrev":hw_p,"heatRateNorm":hw_n,
        "panAvgNormal":pan_norm,"panDev":pan_dev,"baselineLabel":dates["baseline_label"],
        "numAboveNormal":n_above,
        "avgOnset30Delta":avg_onset30_delta,"avgOnset35Delta":avg_onset35_delta,
        "earlier30":earlier30,"earlier35":earlier35,
        # Aliases for embedded template compatibility
        "avgOnset30Diff":avg_onset30_delta,"avgOnset35Diff":avg_onset35_delta,
        "chartRateCurr":hw_c,"chartRatePrev":hw_p,"chartRateNorm":hw_n,
        "onsetEarlier30":earlier30,"onsetEarlier35":earlier35,
        "cities":json.loads(cj),
    })

    tp = Path(__file__).parent / "summer_monitor_template.html"
    html = tp.read_text(encoding="utf-8") if tp.exists() else get_embedded_template()
    return html.replace("/*__DATA_BLOCK__*/", "const DATA = " + db + ";")

# ═══════════════════════════════════════════════════════════════
# HTML TEMPLATE  
# ═══════════════════════════════════════════════════════════════
def get_embedded_template():
    return r'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Summer Strength Monitor</title>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@600;700;800&family=DM+Sans:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500;600&display=swap" rel="stylesheet">
<style>
:root{--bg:#0C0A09;--bg2:#1C1917;--bg3:#292524;--bg4:#44403C;--bd:#44403C;--bd2:#292524;--t1:#FAFAF9;--t2:#A8A29E;--t3:#78716C;--a3:#FCD34D;--a4:#FBBF24;--a5:#F59E0B;--a6:#D97706;--o4:#FB923C;--o5:#F97316;--r4:#F87171;--r5:#EF4444;--g4:#4ADE80;--b4:#60A5FA;--cy:#22D3EE;}
*{margin:0;padding:0;box-sizing:border-box;}
body{font-family:'DM Sans',sans-serif;background:var(--bg);color:var(--t1);min-height:100vh;}
.noise{position:fixed;top:0;left:0;width:100%;height:100%;background-image:url("data:image/svg+xml,%3Csvg viewBox='0 0 256 256' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.9' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)' opacity='0.03'/%3E%3C/svg%3E");pointer-events:none;z-index:1;}
.glow{position:fixed;top:-200px;right:-200px;width:600px;height:600px;background:radial-gradient(circle,rgba(251,191,36,0.08) 0%,rgba(249,115,22,0.04) 40%,transparent 70%);pointer-events:none;}
.wrap{max-width:1440px;margin:0 auto;padding:24px;position:relative;z-index:2;}
header{display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:20px;flex-wrap:wrap;gap:12px;}
h1{font-family:'Playfair Display',serif;font-size:28px;font-weight:800;background:linear-gradient(135deg,var(--a3),var(--o4));-webkit-background-clip:text;-webkit-text-fill-color:transparent;}
.sub{font-size:11px;color:var(--t3);margin-top:3px;letter-spacing:0.5px;text-transform:uppercase;}
.badges{display:flex;gap:6px;flex-wrap:wrap;align-items:center;}
.badge{display:inline-flex;align-items:center;gap:4px;padding:4px 10px;border-radius:5px;font-size:11px;font-weight:500;background:var(--bg3);border:1px solid var(--bd);color:var(--t2);}
.badge.live{border-color:var(--a6);color:var(--a4);}
.badge.live::before{content:'';width:6px;height:6px;border-radius:50%;background:var(--a4);animation:pulse 2s infinite;}
@keyframes pulse{0%,100%{opacity:1}50%{opacity:0.4}}
.verdict{padding:16px 20px;border-radius:10px;margin-bottom:20px;background:linear-gradient(135deg,rgba(251,191,36,0.1),rgba(239,68,68,0.06));border:1px solid rgba(251,191,36,0.2);}
.verdict h2{font-family:'Playfair Display',serif;font-size:20px;margin-bottom:6px;}
.verdict p{font-size:12px;color:var(--t2);line-height:1.6;}
.ssi-block{float:right;text-align:right;margin-left:16px;}
.ssi-big{font-family:'JetBrains Mono',monospace;font-size:44px;font-weight:700;line-height:1;}
.ssi-label{font-size:9px;text-transform:uppercase;letter-spacing:1px;color:var(--t3);}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(165px,1fr));gap:8px;margin-bottom:20px;}
.card{padding:14px;border-radius:8px;background:var(--bg2);border:1px solid var(--bd2);}
.card .lbl{font-size:9px;text-transform:uppercase;letter-spacing:0.7px;color:var(--t3);font-weight:600;margin-bottom:5px;}
.card .val{font-family:'JetBrains Mono',monospace;font-size:22px;font-weight:600;}
.card .dt{font-family:'JetBrains Mono',monospace;font-size:10px;font-weight:500;margin-top:3px;padding:2px 5px;border-radius:3px;display:inline-block;}
.dh{background:rgba(239,68,68,0.15);color:var(--r4);}.dc{background:rgba(59,130,246,0.15);color:var(--b4);}.dn{background:rgba(168,162,158,0.15);color:var(--t2);}
.section{background:var(--bg2);border:1px solid var(--bd2);border-radius:10px;padding:16px;margin-bottom:16px;}
.section h3{font-family:'Playfair Display',serif;font-size:16px;font-weight:700;margin-bottom:12px;}
.section-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px;flex-wrap:wrap;gap:8px;}
.chart-container{position:relative;width:100%;height:280px;}.chart-container canvas{width:100%!important;height:100%!important;}
.chart-legend{display:flex;gap:12px;font-size:11px;flex-wrap:wrap;}.chart-legend span{display:flex;align-items:center;gap:4px;color:var(--t2);}.chart-legend .dot{width:7px;height:7px;border-radius:50%;}
.view-controls{display:flex;gap:2px;background:var(--bg3);padding:2px;border-radius:5px;border:1px solid var(--bd2);flex-wrap:wrap;}
.vbtn{padding:4px 10px;border-radius:3px;border:none;background:transparent;color:var(--t3);font-family:'DM Sans',sans-serif;font-size:10px;font-weight:500;cursor:pointer;transition:all 0.2s;}
.vbtn.active{background:var(--a6);color:#FFFBEB;}.vbtn:hover:not(.active){color:var(--t1);}
.tbl-wrap{overflow-x:auto;}
table{width:100%;border-collapse:collapse;font-size:11px;}
th{text-align:left;padding:6px 10px;font-size:8px;text-transform:uppercase;letter-spacing:0.6px;color:var(--t3);font-weight:600;border-bottom:1px solid var(--bd2);background:var(--bg3);white-space:nowrap;position:sticky;top:0;}
td{padding:6px 10px;border-bottom:1px solid var(--bd2);font-family:'JetBrains Mono',monospace;font-size:10px;white-space:nowrap;}
td.city{font-family:'DM Sans',sans-serif;font-weight:500;font-size:11px;}
tr:hover td{background:rgba(251,191,36,0.03);}
.signal{padding:2px 6px;border-radius:3px;font-size:8px;font-weight:700;letter-spacing:0.4px;}
.sig-s{background:rgba(239,68,68,0.15);color:var(--r4);}.sig-i{background:rgba(168,162,158,0.1);color:var(--t3);}.sig-w{background:rgba(59,130,246,0.15);color:var(--b4);}
.region-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(240px,1fr));gap:8px;}
.region{padding:12px;border-radius:8px;background:var(--bg2);border:1px solid var(--bd2);}
.region .rn{font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:0.6px;color:var(--t2);margin-bottom:8px;padding-bottom:5px;border-bottom:1px solid var(--bd2);}
.region .row{display:flex;justify-content:space-between;padding:3px 0;font-size:11px;}.region .cn{color:var(--t2);}.region .tv{font-family:'JetBrains Mono',monospace;font-weight:500;}
.method{margin-top:10px;padding:12px;border-radius:7px;background:var(--bg3);font-size:10px;color:var(--t3);line-height:1.5;}.method strong{color:var(--t2);}.method code{color:var(--a4);}
.footer{text-align:center;padding:16px 0;font-size:9px;color:var(--t3);border-top:1px solid var(--bd2);margin-top:10px;}
/* Heatmap */
.heatmap-grid{display:grid;grid-template-columns:120px repeat(6,1fr);gap:2px;font-size:10px;}
.hm-header{padding:6px 4px;font-weight:600;color:var(--t3);text-transform:uppercase;font-size:8px;letter-spacing:0.5px;text-align:center;}
.hm-city{padding:6px 8px;color:var(--t2);font-weight:500;display:flex;align-items:center;}
.hm-cell{padding:6px 4px;text-align:center;border-radius:3px;font-family:'JetBrains Mono',monospace;font-weight:600;font-size:10px;}
/* Onset table */
.onset-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(300px,1fr));gap:8px;}
.onset-card{padding:12px;border-radius:8px;background:var(--bg3);border:1px solid var(--bd2);}
.onset-card h4{font-size:12px;font-weight:600;margin-bottom:8px;color:var(--t2);}
.onset-row{display:flex;justify-content:space-between;padding:4px 0;font-size:11px;border-bottom:1px solid var(--bd2);}
.onset-row:last-child{border:none;}
@media(max-width:768px){.wrap{padding:14px;}h1{font-size:22px;}.card .val{font-size:18px;}.ssi-block{float:none;text-align:center;margin:0 0 10px 0;}.chart-container{height:200px;}.heatmap-grid{font-size:8px;}}
</style>
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
</head>
<body>
<div class="noise"></div><div class="glow"></div>
<div class="wrap">
  <header><div><h1>☀ Summer Strength Monitor</h1><div class="sub">India Consumer Equities — Full Jan-Jun Pre-Monsoon Tracker</div></div><div style="display:flex;gap:8px;align-items:center;flex-wrap:wrap;"><div class="badges" id="badges"></div><button onclick="exportExcel()" style="padding:6px 14px;border-radius:6px;border:1px solid var(--a6);background:var(--bg3);color:var(--a4);font-family:'DM Sans',sans-serif;font-size:11px;font-weight:600;cursor:pointer;">📥 Download Excel</button></div></header>
  <div class="verdict" id="verdict"></div>
  <div class="cards" id="cards"></div>

  <!-- Main Chart -->
  <div class="section" id="chartSection">
    <div class="section-header">
      <h3 id="chartTitle">Daily Max Temperature — Pan-India Average</h3>
      <div style="display:flex;gap:10px;align-items:center;flex-wrap:wrap;">
        <div class="chart-legend" id="legend"></div>
        <div class="view-controls">
          <button class="vbtn active" onclick="switchChart('avg',this)">Temp</button>
          <button class="vbtn" onclick="switchChart('hot35',this)">Cities≥35°</button>
          <button class="vbtn" onclick="switchChart('cumhot',this)">Cumul Hot</button>
          <button class="vbtn" onclick="switchChart('heatrate',this)">Heat Rate</button>
          <button class="vbtn" onclick="switchChart('rain',this)">Rain</button>
        </div>
      </div>
    </div>
    <div class="chart-container"><canvas id="mainChart"></canvas></div>
  </div>

  <!-- Summer Onset Tracker -->
  <div class="section"><h3>Summer Onset Tracker</h3><div id="onsetSection"></div></div>

  <!-- Monthly Heatmap -->
  <div class="section"><h3>Monthly Deviation from 5yr Normal (°C)</h3><div class="heatmap-grid" id="heatmap"></div></div>

  <!-- City Table -->
  <div class="section">
    <div class="section-header">
      <h3>City-Level Comparison (Mar-Jun Metrics)</h3>
      <div class="view-controls">
        <button class="vbtn active" onclick="sortBy('delta',this)">YoY Δ</button>
        <button class="vbtn" onclick="sortBy('dev',this)">vs Normal</button>
        <button class="vbtn" onclick="sortBy('onset',this)">Onset</button>
        <button class="vbtn" onclick="sortBy('hot',this)">Hot Days</button>
        <button class="vbtn" onclick="sortBy('ssi',this)">SSI</button>
      </div>
    </div>
    <div class="tbl-wrap"><table><thead id="thead"></thead><tbody id="cityTable"></tbody></table></div>
  </div>

  <!-- Regions -->
  <div class="section"><h3>Regional Assessment</h3><div class="region-grid" id="regionGrid"></div></div>

    <div class="method">
      <strong>Window:</strong> Jan 1 – Jun 30 (full pre-monsoon). SSI computed on Mar-Jun data; onset/heating rate use Jan-Jun.<br>
      <strong>Onset Tracker:</strong> First date max temp crosses 30°C (fan/cooler consideration) and 35°C (impulse beverage). Compared YoY.<br>
      <strong>Heating Rate:</strong> Weekly avg temp change — steeper climb = faster onset. Green dashed = 5yr normal rate.<br>
      <strong>Cumul Hot Days:</strong> Running total of city-days ≥30°C (cyan) + ≥35°C (amber). Directly drives beverage/cooling demand.<br>
      <strong>Heatmap:</strong> Monthly avg temp deviation from 5yr normal per city. Red ≥ +3°C, orange ≥ +1.5°C, blue ≤ -1.5°C.<br>
      <strong>Data:</strong> Open-Meteo (ERA5 + hi-res). <strong>Refresh:</strong> <code>python summer_monitor.py</code>
    </div>
  </div>

  <!-- Methodology & How to Analyze -->
  <div class="section">
    <h3 style="cursor:pointer" onclick="document.getElementById('methBody').style.display=document.getElementById('methBody').style.display==='none'?'block':'none'">
      📖 How This Dashboard Works <span style="font-size:12px;color:var(--t3);font-weight:400">(click to expand/collapse)</span>
    </h3>
    <div id="methBody" style="display:none;font-size:12px;color:var(--t2);line-height:1.8;">

      <div style="margin-bottom:20px;padding:14px;border-radius:8px;background:var(--bg3);border-left:3px solid var(--a4)">
        <strong style="color:var(--a4);font-size:13px">Quick Guide: How to Read This Dashboard</strong><br>
        ① Check the <strong>SSI</strong> in the header — above 1.05 = stronger summer YoY (bullish VBL/Dabur/Havells), below 0.95 = weaker.<br>
        ② Check <strong>"vs 5yr Normal"</strong> — even if SSI is in-line, a +2°C deviation means absolute heat levels are elevated (supports estimates).<br>
        ③ Check <strong>Cumulative Hot Days</strong> chart — if the 2026 line is above 2025 at the same point, more consumers across more cities are in heat-driven buying mode.<br>
        ④ Check <strong>Onset Tracker</strong> — negative days = summer arrived earlier = more selling days in the quarter.
      </div>

      <strong style="color:var(--t1);font-size:13px">━━ DATA SOURCES ━━</strong><br><br>
      <strong>Open-Meteo (29 cities):</strong> Uses ERA5 reanalysis + high-resolution national weather models. Grid resolution ~14km.
      Data is gridded (not station-level), so readings may be 1-2°C below IMD station data for inland cities.
      All YoY comparisons use the same grid cell, so relative signals are internally consistent.<br><br>
      <strong>Visual Crossing — Mumbai only:</strong> Uses actual METAR station data from Santa Cruz airport (IMD station 43003).
      This is the same data IMD publishes. Open-Meteo's grid cannot resolve Mumbai's narrow 3km-wide peninsula —
      every grid cell either blends with the Arabian Sea (reads 28-31°C) or falls in inland Thane (reads 37-40°C).
      VC solves this with direct station observation data.<br><br>
      <strong>5yr Baseline:</strong> Daily normals computed from 2020-2024 data (Open-Meteo) or 2022-2024 (Mumbai/VC).
      Per-city, per-day-of-season average. Used for "vs Normal" deviation and the green dashed line on charts.<br><br>

      <strong style="color:var(--t1);font-size:13px">━━ SEASON WINDOW ━━</strong><br><br>
      <strong>Jan 1 – Jun 30</strong> is the full tracking window. Data is fetched for this entire period for both current and previous year.<br>
      <strong>Summary cards, city table, regional assessment, and SSI</strong> use <strong>Mar 1 onwards only</strong> — because Jan-Feb temperatures
      (15-25°C in North India) would dilute the summer signal. These sections reflect actual summer intensity.<br>
      <strong>Charts, onset tracker, heating rate, and monthly heatmap</strong> use the <strong>full Jan-Jun window</strong> — because they're
      designed to show how summer builds up, including the winter-to-summer transition that drives onset timing.<br><br>

      <strong style="color:var(--t1);font-size:13px">━━ SSI (SUMMER STRENGTH INDEX) ━━</strong><br><br>
      A <strong>YoY ratio</strong> measuring whether this summer is stronger or weaker than last year. Computed per city on Mar-Jun data, then averaged across all cities.<br><br>
      <strong>Formula:</strong> SSI = (Temp Score × 50%) + (Hot Days Score × 30%) + (Rain Score × 20%)<br><br>
      <table style="width:100%;font-size:11px;border-collapse:collapse;margin:8px 0;">
        <tr style="border-bottom:1px solid var(--bd2)">
          <td style="padding:6px;color:var(--a4);font-weight:600;width:25%">Temp Score (50%)</td>
          <td style="padding:6px">Mar avg max 2026 ÷ Mar avg max 2025.<br>Example: 34.4°C / 31.0°C = 1.11 (11% hotter)</td>
        </tr>
        <tr style="border-bottom:1px solid var(--bd2)">
          <td style="padding:6px;color:var(--a4);font-weight:600">Hot Days Score (30%)</td>
          <td style="padding:6px">Days ≥35°C ratio (2026 / 2025), capped between 0.5x and 2.0x.<br>
          When base year has <5 hot days, uses dampened absolute difference (+0.05 per extra day, capped at 1.3x).<br>
          This prevents small-number distortion — e.g. Mumbai going from 2→7 hot days doesn't overwhelm the index.</td>
        </tr>
        <tr style="border-bottom:1px solid var(--bd2)">
          <td style="padding:6px;color:var(--a4);font-weight:600">Rain Score (20%)</td>
          <td style="padding:6px">Rainfall 2025 ÷ Rainfall 2026 (inverted — less rain = drier = higher score).<br>
          Capped at 1.5x. Dry heat amplifies beverage demand more than humid heat.</td>
        </tr>
      </table>
      <strong>Interpretation:</strong> SSI > 1.05 → <span style="color:var(--r4)">STRONGER</span> (bullish for summer plays) |
      SSI 0.95–1.05 → <span style="color:var(--t3)">IN-LINE</span> |
      SSI < 0.95 → <span style="color:var(--b4)">WEAKER</span> (cautious)<br><br>
      <strong>Important:</strong> SSI is a relative (YoY) measure. A city can show SSI = 0.95 (WEAKER) even if it's 4°C above historical normal —
      because last year was even hotter. Always read SSI alongside "vs Normal" for the complete picture.<br><br>

      <strong style="color:var(--t1);font-size:13px">━━ vs 5yr NORMAL ━━</strong><br><br>
      Shows how current temperatures compare to the 2020-2024 average for the same calendar period.
      This is the same framework IMD uses in their bulletins (e.g. "7.6°C above normal").<br>
      <span style="color:var(--r4)">WELL ABOVE (≥3°C)</span> |
      <span style="color:var(--o5)">ABOVE (≥1.5°C)</span> |
      <span style="color:var(--t3)">NEAR NORMAL (±1.5°C)</span> |
      <span style="color:var(--b4)">BELOW (≤-1.5°C)</span><br><br>
      <strong>For your thesis:</strong> Even when SSI is in-line (~1.0), a high deviation from normal (e.g. +2°C) means both this year
      AND last year are running well above historical levels — the elevated demand base is structurally sustained.<br><br>

      <strong style="color:var(--t1);font-size:13px">━━ ONSET TRACKER ━━</strong><br><br>
      Tracks the <strong>first date</strong> each city's max temperature crosses two thresholds:<br>
      <strong>≥30°C</strong> — Fan/cooler consideration starts. AC showroom footfall picks up. Consumer durables lead indicator.<br>
      <strong>≥35°C</strong> — Impulse beverage territory. Non-linear demand uplift for VBL (cold drinks), Dabur (glucose, Real juices),
      ice cream (HUL Kwality Walls). This is the threshold that drives quarterly volume surprises.<br><br>
      <strong>Delta:</strong> Negative = 2026 crossed earlier than 2025 (green, bullish — more selling days).<br>
      Positive = 2026 crossed later (red — fewer selling days so far).<br>
      <strong>Example:</strong> If Delhi crosses 35°C on Mar 7 (2026) vs Mar 25 (2025) = <span style="color:var(--g4)">18d earlier</span>.
      At VBL's daily run-rate, that's 18 extra days of peak demand in Q1 FY27.<br><br>

      <strong style="color:var(--t1);font-size:13px">━━ CHARTS ━━</strong><br><br>
      <strong>Avg Temp:</strong> Daily pan-India average max temperature. Amber = 2026, grey dashed = 2025 full season, green dashed = 5yr normal.
      The gap between amber and green is your deviation-from-normal.<br>
      <strong>Cities ≥35°C:</strong> How many of the 30 cities crossed 35°C each day. Higher = broader geographic heat spread.<br>
      <strong>Cumulative Hot Days:</strong> Running total of city-days. Amber = ≥30°C (warm days), Red = ≥35°C (hot days). Solid = 2026, dashed = 2025.
      If the 2026 line is above 2025 at the same point in the season → more heat exposure × population reach → higher beverage/cooling volumes.<br>
      <strong>Heating Rate:</strong> Weekly average max temperature. The slope shows how fast summer is building.
      Steeper than normal (green) = earlier onset. Useful as a February leading indicator.<br>
      <strong>Rainfall:</strong> Lower rainfall during Mar-Jun = drier heat = stronger beverage demand signal. Hot + dry > hot + wet.<br><br>

      <strong style="color:var(--t1);font-size:13px">━━ MONTHLY HEATMAP ━━</strong><br><br>
      Shows each city's deviation from its own 5yr normal for each month (Jan-Jun).
      Color-coded: deep red = well above normal (+4°C+), orange = above (+2°C), grey = normal, blue = below.<br>
      <strong>How to read:</strong> Scan horizontally to see a city's seasonal progression. Scan vertically to see which month was the anomaly.
      If March is deep red across North India but grey in South → the heat belt thesis is concentrated in the right geography for VBL/Dabur.<br><br>

      <span style="color:var(--r4)">HIGH:</span> VBL, Dabur, Havells, Crompton, Voltas, Blue Star — revenues directly driven by summer intensity. Q1 FY27 (Apr-Jun) is 35-40% of annual volumes for VBL.<br>
      <span style="color:var(--a4)">MEDIUM:</span> HUL (Kwality Walls, beverages mix shift), Marico (hair oils seasonal) — meaningful but not dominant seasonal driver.<br>
      <span style="color:var(--b4)">LOW:</span> Nestlé, Titan, DMart, Britannia — minor seasonal mix impact. Not a primary thesis driver.<br><br>

      <strong style="color:var(--t1);font-size:13px">━━ KEY RISKS & CAVEATS ━━</strong><br><br>
      • <strong>Western disturbances</strong> can bring temporary relief (5-10°C drop for 2-3 days), pulling down weekly averages. This is noise, not signal — look at the trend, not individual days.<br>
      • <strong>Grid resolution bias:</strong> Open-Meteo reads 1-2°C below IMD station data due to grid averaging. All comparisons are internally consistent (same grid cell both years), so relative signals are valid.<br>
      • <strong>Extreme heat → drought risk:</strong> If heat persists without rain, rural demand can be hurt (negative for FMCG distribution). Monitor rainfall chart.<br>
      • <strong>Crude oil / LPG prices:</strong> Iran conflict → oil spike → disposable income squeeze could offset volume gains from heat. Consider macro overlay.<br>
      • <strong>SSI stabilizes by late April</strong> — with 50+ days of data, the index becomes reliable. Early March readings (21 days) can be noisy.<br><br>

      <strong style="color:var(--t1);font-size:13px">━━ HOW TO REFRESH ━━</strong><br><br>
      Run <code style="color:var(--a4)">python summer_monitor.py</code> in CMD. Takes ~90 seconds for 30 cities.<br>
      Opens the updated HTML automatically. Re-run weekly for tracking, or daily during critical heatwave periods.<br>
      Validate: <code style="color:var(--a4)">python summer_monitor.py --validate</code> shows resolved coordinates and cross-check URLs.

    </div>
  </div>

  <div class="footer" id="footer"></div>
</div>

<script>
/*__DATA_BLOCK__*/
const D=DATA;
let currentView='avg',currentSort='delta';
function tc(t){return t>=40?'var(--r5)':t>=38?'var(--r4)':t>=36?'var(--o5)':t>=34?'var(--a4)':'var(--t2)';}
function sign(v){return v>=0?'+'+v:''+v;}
function devColor(d){return d>=3?'var(--r4)':d>=1.5?'var(--o5)':d<=-1.5?'var(--b4)':'var(--t2)';}
function hmBg(d){if(d===undefined||d===null)return'var(--bg3)';if(d>=4)return'rgba(239,68,68,0.35)';if(d>=3)return'rgba(239,68,68,0.25)';if(d>=1.5)return'rgba(251,191,36,0.2)';if(d>=-1.5)return'rgba(168,162,158,0.1)';return'rgba(59,130,246,0.2)';}
function hmColor(d){if(d===undefined||d===null)return'var(--t3)';if(d>=3)return'var(--r4)';if(d>=1.5)return'var(--a4)';if(d<=-1.5)return'var(--b4)';return'var(--t2)';}

// Badges
document.getElementById('badges').innerHTML=`<span class="badge live">Open-Meteo</span><span class="badge">${D.startCurrent} → ${D.endCurrent}</span><span class="badge">${D.numCities} cities · ${D.daysTracked}d</span>`;

// Verdict
document.getElementById('verdict').innerHTML=`
  <div class="ssi-block"><div class="ssi-big" style="color:${D.verdictColor}">${D.panSSI.toFixed(2)}</div><div class="ssi-label">SSI (Mar-Jun YoY)</div></div>
  <h2 style="color:${D.verdictColor}">${D.verdictEmoji} ${D.verdictText}</h2>
  <p>Mar onwards avg max (${D.marDays||21}d): <strong>${D.marAvgCurr}°C</strong> vs <strong>${D.marAvgPrev}°C</strong> (${sign(D.marDelta)}°C YoY).
  vs 5yr normal: <strong>${D.marDev>=0?'+':''}${D.marDev}°C</strong> (${D.marAboveNormal||D.numAboveNormal}/${D.numCities} above normal).
  30°C onset: <strong>${D.avgOnset30Diff>=0?'':'-'}${Math.abs(D.avgOnset30Diff)}d</strong> ${D.avgOnset30Diff<0?'earlier':'later'} (${D.onsetEarlier30}/${D.numCities} cities earlier).
  35°C onset: <strong>${D.avgOnset35Diff>=0?'':'-'}${Math.abs(D.avgOnset35Diff)}d</strong> ${D.avgOnset35Diff<0?'earlier':'later'}.
  ≥35°C city-days (Mar+): <strong>${D.marH35Curr||D.hot35Curr}</strong> vs ${D.marH35Prev||D.hot35Prev} (${sign((D.marH35Curr||D.hot35Curr)-(D.marH35Prev||D.hot35Prev))}).
  → <strong>${D.ssiBias}</strong> for summer-sensitive consumer names.</p>`;

// Cards
const mpd=D.marDelta||D.panDelta, mh35d=(D.marH35Curr||D.hot35Curr)-(D.marH35Prev||D.hot35Prev), mh40d=(D.marH40Curr||D.hot40Curr)-(D.marH40Prev||D.hot40Prev);
document.getElementById('cards').innerHTML=`
  <div class="card"><div class="lbl">Avg Max (Mar onwards, ${D.marDays||21}d)</div><div class="val" style="color:var(--a3)">${D.marAvgCurr||D.panAvgCurr}°C</div><div class="dt ${mpd>0?'dh':mpd<0?'dc':'dn'}">${sign(mpd)}°C vs ${D.lastYear}</div></div>
  <div class="card"><div class="lbl">vs 5yr Normal (Mar+)</div><div class="val" style="color:${devColor(D.marDev||D.panDev)}">${(D.marDev||D.panDev)>=0?'+':''}${D.marDev||D.panDev}°C</div><div class="dt dh">${D.marAboveNormal||D.numAboveNormal}/${D.numCities} above</div></div>
  <div class="card"><div class="lbl">30°C Onset (avg)</div><div class="val">${Math.abs(D.avgOnset30Diff)}d</div><div class="dt ${D.avgOnset30Diff<0?'dh':'dc'}">${D.avgOnset30Diff<0?'Earlier':'Later'} vs ${D.lastYear}</div></div>
  <div class="card"><div class="lbl">35°C Onset (avg)</div><div class="val">${Math.abs(D.avgOnset35Diff)}d</div><div class="dt ${D.avgOnset35Diff<0?'dh':'dc'}">${D.avgOnset35Diff<0?'Earlier':'Later'} vs ${D.lastYear}</div></div>
  <div class="card"><div class="lbl">Hot Days ≥35°C (Mar+)</div><div class="val">${D.marH35Curr||D.hot35Curr}</div><div class="dt ${mh35d>0?'dh':'dc'}">${sign(mh35d)} vs ${D.lastYear}</div></div>
  <div class="card"><div class="lbl">Days ≥40°C (Mar+)</div><div class="val">${D.marH40Curr||D.hot40Curr}</div><div class="dt ${mh40d>0?'dh':'dc'}">${sign(mh40d)}</div></div>
  <div class="card"><div class="lbl">SSI (YoY)</div><div class="val" style="color:${D.verdictColor}">${D.panSSI.toFixed(2)}</div><div class="dt ${D.panSSI>1.05?'dh':D.panSSI<0.95?'dc':'dn'}">${D.ssiSignal}</div></div>`;

// Legend
const legends={
  avg:`<span><span class="dot" style="background:var(--a4)"></span>${D.year}</span><span><span class="dot" style="background:var(--t3)"></span>${D.lastYear}</span><span><span class="dot" style="background:var(--g4);opacity:0.6"></span>5yr Normal</span>`,
  hot35:`<span><span class="dot" style="background:var(--a4)"></span>${D.year}</span><span><span class="dot" style="background:var(--t3)"></span>${D.lastYear}</span>`,
  cumhot:`<span><span class="dot" style="background:var(--cy)"></span>≥30°C ${D.year}</span><span><span class="dot" style="background:var(--a4)"></span>≥35°C ${D.year}</span><span><span class="dot" style="background:var(--t3)"></span>${D.lastYear} (dashed)</span>`,
  heatrate:`<span><span class="dot" style="background:var(--a4)"></span>${D.year}</span><span><span class="dot" style="background:var(--t3)"></span>${D.lastYear}</span>`,
  rain:`<span><span class="dot" style="background:var(--a4)"></span>${D.year}</span><span><span class="dot" style="background:var(--t3)"></span>${D.lastYear}</span>`,
};
const titles={avg:'Daily Max Temp — Pan-India Avg',hot35:'Cities ≥35°C per Day',cumhot:'Cumulative Hot Days — Pan-India Total',heatrate:'Weekly Heating Rate (°C/week)',rain:'Daily Rainfall — Pan-India Avg'};
document.getElementById('legend').innerHTML=legends.avg;

// Chart data
const CS={
  avg:{curr:D.chartAvgCurr,prev:D.chartAvgPrev,normal:D.chartNormal,label:'°C',dec:1},
  hot35:{curr:D.chartHot35Curr,prev:D.chartHot35Prev,normal:null,label:'Cities ≥35°C',dec:0},
  cumhot:{curr:D.chartCum35Curr,prev:D.chartCum35Prev,normal:null,extra:{curr2:D.chartCum30Curr,prev2:D.chartCum30Prev},label:'City-days',dec:0},
  heatrate:{curr:D.chartRateCurr,prev:D.chartRatePrev,normal:null,label:'°C/week',dec:1,isWeekly:true},
  rain:{curr:D.chartRainCurr,prev:D.chartRainPrev,normal:null,label:'mm',dec:1},
};

function drawChart(view){
  const canvas=document.getElementById('mainChart'),ctx=canvas.getContext('2d');
  const W=canvas.parentElement.clientWidth,H=canvas.parentElement.clientHeight;
  canvas.width=W*2;canvas.height=H*2;ctx.scale(2,2);
  const pad={top:20,right:20,bottom:40,left:50};
  ctx.clearRect(0,0,W,H);
  const d=CS[view],curr=d.curr,prev=d.prev,norm=d.normal||[];
  const extra=d.extra||{};
  const all=[...curr,...prev,...norm,...(extra.curr2||[]),...(extra.prev2||[])].filter(v=>v!==null);
  if(!all.length)return;
  const yMin=Math.floor(Math.min(...all)-(d.dec?2:1)),yMax=Math.ceil(Math.max(...all)+(d.dec?2:1));
  const pW=W-pad.left-pad.right,pH=H-pad.top-pad.bottom;
  const maxD=d.isWeekly?Math.max(curr.length,prev.length):Math.max(curr.length,prev.length,norm.length);
  const xS=i=>pad.left+(i/Math.max(maxD-1,1))*pW;
  const yS=v=>pad.top+pH-((v-yMin)/(yMax-yMin))*pH;

  // Grid
  ctx.strokeStyle='rgba(255,255,255,0.05)';ctx.lineWidth=0.5;
  for(let i=0;i<=5;i++){const val=yMin+(i/5)*(yMax-yMin),y=yS(val);
    ctx.beginPath();ctx.moveTo(pad.left,y);ctx.lineTo(W-pad.right,y);ctx.stroke();
    ctx.fillStyle='rgba(255,255,255,0.3)';ctx.font='10px JetBrains Mono';ctx.textAlign='right';
    ctx.fillText(val.toFixed(d.dec),pad.left-8,y+3);}
  ctx.save();ctx.fillStyle='rgba(255,255,255,0.4)';ctx.font='10px DM Sans';
  ctx.translate(12,pad.top+pH/2);ctx.rotate(-Math.PI/2);ctx.textAlign='center';ctx.fillText(d.label,0,0);ctx.restore();

  // X labels
  ctx.fillStyle='rgba(255,255,255,0.3)';ctx.font='10px JetBrains Mono';ctx.textAlign='center';
  if(d.isWeekly){for(let w=0;w<maxD;w+=4){ctx.fillText('W'+(w+1),xS(w),H-pad.bottom+20);}}
  else{['Jan','Feb','Mar','Apr','May','Jun'].forEach((m,idx)=>{const day=[0,31,59,90,120,151][idx];if(day<maxD)ctx.fillText(m,xS(day),H-pad.bottom+20);});}

  function drawLine(arr,color,width,dash){
    if(!arr||!arr.length)return;
    ctx.strokeStyle=color;ctx.lineWidth=width;if(dash)ctx.setLineDash(dash);
    ctx.beginPath();arr.forEach((v,i)=>{if(v===null)return;const x=xS(i),y=yS(v);i===0||arr[i-1]===null?ctx.moveTo(x,y):ctx.lineTo(x,y);});ctx.stroke();if(dash)ctx.setLineDash([]);}

  // Draw order: normal, extra prev, prev, extra curr, curr
  if(norm.length)drawLine(norm,'rgba(74,222,128,0.4)',2,[6,3]);
  if(extra.prev2)drawLine(extra.prev2,'rgba(34,211,238,0.25)',1,[3,3]);
  if(extra.curr2){drawLine(extra.curr2,'rgba(34,211,238,0.6)',2,null);}
  drawLine(prev,'rgba(120,113,108,0.5)',1.5,[4,4]);

  // Current with fill
  if(curr.length){
    const gr=ctx.createLinearGradient(0,pad.top,0,pad.top+pH);
    gr.addColorStop(0,'rgba(251,191,36,0.12)');gr.addColorStop(1,'rgba(251,191,36,0)');
    ctx.fillStyle=gr;ctx.beginPath();ctx.moveTo(xS(0),pad.top+pH);
    curr.forEach((v,i)=>{if(v!==null)ctx.lineTo(xS(i),yS(v));});
    ctx.lineTo(xS(curr.length-1),pad.top+pH);ctx.closePath();ctx.fill();
    drawLine(curr,'#FBBF24',2.5,null);
    const last=curr.length-1;
    if(curr[last]!==null){ctx.beginPath();ctx.arc(xS(last),yS(curr[last]),4,0,Math.PI*2);ctx.fillStyle='#FBBF24';ctx.fill();ctx.strokeStyle='#0C0A09';ctx.lineWidth=2;ctx.stroke();}
    if(curr.length<maxD){const x=xS(curr.length-1);ctx.strokeStyle='rgba(251,191,36,0.3)';ctx.lineWidth=1;ctx.setLineDash([2,3]);ctx.beginPath();ctx.moveTo(x,pad.top);ctx.lineTo(x,pad.top+pH);ctx.stroke();ctx.setLineDash([]);ctx.fillStyle='rgba(251,191,36,0.6)';ctx.font='9px DM Sans';ctx.textAlign='center';ctx.fillText('Today',x,pad.top-6);}
  }
}

function switchChart(v,btn){currentView=v;
  document.querySelectorAll('#chartSection .vbtn').forEach(b=>b.classList.remove('active'));
  btn.classList.add('active');document.getElementById('chartTitle').textContent=titles[v]||'';
  document.getElementById('legend').innerHTML=legends[v]||legends.avg;drawChart(v);}

// ── Summer Onset Tracker ──
function renderOnset(){
  const sorted30=[...D.cities].sort((a,b)=>(a.onset30diff||999)-(b.onset30diff||999));
  const sorted35=[...D.cities].sort((a,b)=>(a.onset35diff||999)-(b.onset35diff||999));
  document.getElementById('onsetSection').innerHTML=`
  <div class="onset-grid">
    <div class="onset-card"><h4>First ≥30°C (Fan/Cooler Season Start)</h4>
      ${sorted30.map(c=>{const diff=c.onset30diff;const dc=diff<0?'var(--g4)':diff>0?'var(--r4)':'var(--t3)';
        return `<div class="onset-row"><span style="color:var(--t2)">${c.city}</span><span style="font-family:'JetBrains Mono';font-size:10px;"><span style="color:var(--a4)">${c.onset30c?c.onset30c.slice(5):'—'}</span> <span style="color:var(--t3)">vs ${c.onset30p?c.onset30p.slice(5):'—'}</span> <span style="color:${dc};font-weight:600">${diff!==null?(diff<0?Math.abs(diff)+'d earlier':diff>0?diff+'d later':'same'):'—'}</span></span></div>`;}).join('')}
    </div>
    <div class="onset-card"><h4>First ≥35°C (Impulse Beverage Threshold)</h4>
      ${sorted35.map(c=>{const diff=c.onset35diff;const dc=diff<0?'var(--g4)':diff>0?'var(--r4)':'var(--t3)';
        return `<div class="onset-row"><span style="color:var(--t2)">${c.city}</span><span style="font-family:'JetBrains Mono';font-size:10px;"><span style="color:var(--r4)">${c.onset35c?c.onset35c.slice(5):'—'}</span> <span style="color:var(--t3)">vs ${c.onset35p?c.onset35p.slice(5):'—'}</span> <span style="color:${dc};font-weight:600">${diff!==null?(diff<0?Math.abs(diff)+'d earlier':diff>0?diff+'d later':'same'):'—'}</span></span></div>`;}).join('')}
    </div>
  </div>`;
}

// ── Monthly Heatmap ──
function renderHeatmap(){
  const months=['Jan','Feb','Mar','Apr','May','Jun'];
  let html='<div class="hm-header"></div>';
  months.forEach(m=>{html+=`<div class="hm-header">${m}</div>`;});
  D.cities.forEach(c=>{
    html+=`<div class="hm-city">${c.city}</div>`;
    for(let m=1;m<=6;m++){
      const d=c.monthlyDev[m];
      html+=`<div class="hm-cell" style="background:${hmBg(d)};color:${hmColor(d)}">${d!==undefined&&d!==null?(d>=0?'+':'')+d.toFixed(1):'—'}</div>`;
    }
  });
  document.getElementById('heatmap').innerHTML=html;
}

// ── City Table ──
document.getElementById('thead').innerHTML=`<tr><th>City</th><th>Region</th><th>Avg Max</th><th>vs ${D.lastYear}</th><th>vs Normal</th><th>30°C On</th><th>35°C On</th><th>≥35°C</th><th>≥40°C</th><th>SSI</th><th>Signal</th></tr>`;

function renderTable(){
  let s=[...D.cities];
  if(currentSort==='delta')s.sort((a,b)=>(b.marDelta||b.delta)-(a.marDelta||a.delta));
  else if(currentSort==='dev')s.sort((a,b)=>(b.marDev||b.devFromNormal||0)-(a.marDev||a.devFromNormal||0));
  else if(currentSort==='onset')s.sort((a,b)=>(a.onset35diff||999)-(b.onset35diff||999));
  else if(currentSort==='hot')s.sort((a,b)=>b.hot35Curr-a.hot35Curr);
  else if(currentSort==='ssi')s.sort((a,b)=>b.ssi-a.ssi);
  document.getElementById('cityTable').innerHTML=s.map(c=>{
    const sc=c.signal==='STRONGER'?'sig-s':c.signal==='WEAKER'?'sig-w':'sig-i';
    const d35=c.onset35diff;
    return `<tr>
      <td class="city">${c.city}</td>
      <td style="color:var(--t3);font-family:'DM Sans';font-size:9px">${c.region}</td>
      <td style="color:${tc(c.marAvgCurr||c.avgMaxCurr)};font-weight:600">${(c.marAvgCurr||c.avgMaxCurr)}°C</td>
      <td style="color:${(c.marDelta||c.delta)>=0?'var(--r4)':'var(--b4)'};font-weight:600">${(c.marDelta||c.delta)>=0?'+':''}${(c.marDelta||c.delta)}°C</td>
      <td style="color:${devColor((c.marDev!==undefined&&c.marDev!==null)?c.marDev:(c.devFromNormal||0))}">${(c.marDev!==undefined&&c.marDev!==null)?(c.marDev>=0?'+':'')+c.marDev+'°C':(c.devFromNormal!==null?(c.devFromNormal>=0?'+':'')+c.devFromNormal+'°C':'—')}</td>
      <td style="color:var(--a4)">${c.onset30c?c.onset30c.slice(5):'—'}</td>
      <td style="color:var(--r4)">${c.onset35c?c.onset35c.slice(5):'—'} <span style="color:${d35<0?'var(--g4)':d35>0?'var(--r5)':'var(--t3)'};font-size:9px">${d35!==null?(d35<0?Math.abs(d35)+'d↑':d35>0?d35+'d↓':'='):'—'}</span></td>
      <td>${c.hot35Curr} <span style="color:var(--t3);font-size:9px">vs ${c.hot35Prev}</span></td>
      <td style="color:${c.hot40Curr>0?'var(--r4)':'var(--t3)'}">${c.hot40Curr}</td>
      <td style="color:${c.ssi>1.05?'var(--r4)':c.ssi<0.95?'var(--b4)':'var(--t2)'};font-weight:600">${c.ssi.toFixed(2)}</td>
      <td><span class="signal ${sc}">${c.signal}</span></td></tr>`;
  }).join('');
}

function sortBy(s,btn){currentSort=s;btn.parentElement.querySelectorAll('.vbtn').forEach(b=>b.classList.remove('active'));btn.classList.add('active');renderTable();}

// ── Regions ──
function renderRegions(){
  const R={};D.cities.forEach(c=>{if(!R[c.region])R[c.region]=[];R[c.region].push(c);});
  document.getElementById('regionGrid').innerHTML=Object.entries(R).map(([n,cs])=>{
    const ad=(cs.reduce((s,c)=>s+(c.marDelta||c.delta||0),0)/cs.length).toFixed(1);
    const devs=cs.filter(c=>(c.marDev||c.devFromNormal)!==null&&(c.marDev||c.devFromNormal)!==undefined);
    const avgDev=devs.length?(devs.reduce((s,c)=>s+(c.marDev||c.devFromNormal||0),0)/devs.length).toFixed(1):'—';
    return `<div class="region"><div class="rn">${n}<span style="float:right;font-family:'JetBrains Mono';font-size:10px"><span style="color:${parseFloat(ad)>=0?'var(--r4)':'var(--b4)'}">${parseFloat(ad)>=0?'+':''}${ad}° YoY</span> · <span style="color:${devColor(parseFloat(avgDev))}">${parseFloat(avgDev)>=0?'+':''}${avgDev}° v.n</span></span></div>
    ${cs.map(c=>`<div class="row"><span class="cn">${c.city}</span><span><span class="tv" style="color:${tc(c.marAvgCurr||c.avgMaxCurr)}">${(c.marAvgCurr||c.avgMaxCurr)}°C</span><span style="color:${(c.marDelta||c.delta)>=0?'var(--r4)':'var(--b4)'};font-family:'JetBrains Mono';font-size:9px;margin-left:4px">${(c.marDelta||c.delta)>=0?'+':''}${(c.marDelta||c.delta)}°</span></span></div>`).join('')}</div>`;
  }).join('');
}


// Footer
document.getElementById('footer').innerHTML=`Summer Strength Monitor v4 — Generated ${D.generatedAt} — Open-Meteo (CC BY 4.0) + Visual Crossing (Mumbai METAR) — Axis Capital Consumer Research`;

// ── Excel Export ──
function exportExcel(){
  const wb = XLSX.utils.book_new();
  const yr=D.year, ly=D.lastYear;

  // Sheet 1: Summary
  const sumRows = [
    ['Summer Strength Monitor — Summary', '','',''],
    ['Generated', D.generatedAt,'',''],
    ['Period', D.startCurrent + ' to ' + D.endCurrent,'',''],
    ['Cities', D.numCities,'',''],
    ['','','',''],
    ['Metric (Mar onwards)', yr, ly, 'Delta'],
    ['Avg Max Temp (°C)', D.marAvgCurr, D.marAvgPrev, D.marDelta],
    ['vs 5yr Normal (°C)', D.marDev||D.panDev, '', ''],
    ['Cities Above Normal', D.marAboveNormal||D.numAboveNormal, '', ''],
    ['Hot Days ≥35°C (city-days)', D.marH35Curr||D.hot35Curr, D.marH35Prev||D.hot35Prev, (D.marH35Curr||D.hot35Curr)-(D.marH35Prev||D.hot35Prev)],
    ['Days ≥40°C', D.marH40Curr||D.hot40Curr, D.marH40Prev||D.hot40Prev, ''],
    ['SSI (YoY)', D.panSSI, '', D.ssiSignal],
    ['30°C Onset Delta (avg days)', D.avgOnset30Diff, '', D.avgOnset30Diff<0?'Earlier':'Later'],
    ['35°C Onset Delta (avg days)', D.avgOnset35Diff, '', D.avgOnset35Diff<0?'Earlier':'Later'],
  ];
  const ws1 = XLSX.utils.aoa_to_sheet(sumRows);
  ws1['!cols'] = [{wch:30},{wch:12},{wch:12},{wch:12}];
  XLSX.utils.book_append_sheet(wb, ws1, 'Summary');

  // Sheet 2: City Data (Mar onwards)
  const cityHeader = ['City','Region','Avg Max '+yr,'Avg Max '+ly,'YoY Delta','vs Normal','5yr Normal',
    '30°C Onset '+yr,'30°C Onset '+ly,'30°C Delta (days)',
    '35°C Onset '+yr,'35°C Onset '+ly,'35°C Delta (days)',
    'Hot Days ≥35 ('+yr+')','Hot Days ≥35 ('+ly+')','Days ≥40','SSI','Signal'];
  const cityRows = D.cities.map(c=>[
    c.city, c.region,
    c.marAvgCurr||c.avgMaxCurr, c.marAvgPrev||c.avgMaxPrev, c.marDelta||c.delta,
    c.marDev!==undefined&&c.marDev!==null?c.marDev:(c.devFromNormal||''),
    c.marAvgNormal||c.avgNormal||'',
    c.onset30c||'', c.onset30p||'', c.onset30diff!==null?c.onset30diff:'',
    c.onset35c||'', c.onset35p||'', c.onset35diff!==null?c.onset35diff:'',
    c.hot35Curr, c.hot35Prev, c.hot40Curr, c.ssi, c.signal
  ]);
  const ws2 = XLSX.utils.aoa_to_sheet([cityHeader, ...cityRows]);
  ws2['!cols'] = [{wch:16},{wch:8},{wch:10},{wch:10},{wch:9},{wch:9},{wch:10},{wch:12},{wch:12},{wch:12},{wch:12},{wch:12},{wch:12},{wch:10},{wch:10},{wch:7},{wch:6},{wch:10}];
  XLSX.utils.book_append_sheet(wb, ws2, 'City Data');

  // Sheet 3: Monthly Heatmap (deviation from normal)
  const mths = ['Jan','Feb','Mar','Apr','May','Jun'];
  const hmHeader = ['City','Region',...mths.map(m=>m+' Avg °C'),...mths.map(m=>m+' Normal °C'),...mths.map(m=>m+' Dev °C')];
  const hmRows = D.cities.map(c=>{
    const row = [c.city, c.region];
    for(let m=1;m<=6;m++) row.push(c.monthlyCurr&&c.monthlyCurr[m]!==undefined?c.monthlyCurr[m]:'');
    for(let m=1;m<=6;m++) row.push(c.monthlyNorm&&c.monthlyNorm[m]!==undefined?c.monthlyNorm[m]:'');
    for(let m=1;m<=6;m++) row.push(c.monthlyDev&&c.monthlyDev[m]!==undefined?c.monthlyDev[m]:'');
    return row;
  });
  const ws3 = XLSX.utils.aoa_to_sheet([hmHeader, ...hmRows]);
  XLSX.utils.book_append_sheet(wb, ws3, 'Monthly Heatmap');

  // Sheet 4: Daily Pan-India (chart data)
  const maxLen = Math.max(D.chartAvgCurr.length, D.chartAvgPrev.length, D.chartNormal.length);
  const dailyHeader = ['Day of Season','Avg Max '+yr,'Avg Max '+ly,'5yr Normal','Cities ≥35 '+yr,'Cities ≥35 '+ly,
    'Cum ≥30 '+yr,'Cum ≥35 '+yr,'Cum ≥30 '+ly,'Cum ≥35 '+ly,'Rain '+yr,'Rain '+ly];
  const dailyRows = [];
  for(let i=0;i<maxLen;i++){
    dailyRows.push([i+1,
      i<D.chartAvgCurr.length?D.chartAvgCurr[i]:'',
      i<D.chartAvgPrev.length?D.chartAvgPrev[i]:'',
      i<D.chartNormal.length?D.chartNormal[i]:'',
      D.chartHot35Curr&&i<D.chartHot35Curr.length?D.chartHot35Curr[i]:'',
      i<(D.chartHot35Prev||D.chartHotPrev||[]).length?(D.chartHot35Prev||D.chartHotPrev)[i]:'',
      i<D.chartCum30Curr.length?D.chartCum30Curr[i]:'',
      i<D.chartCum35Curr.length?D.chartCum35Curr[i]:'',
      i<D.chartCum30Prev.length?D.chartCum30Prev[i]:'',
      i<D.chartCum35Prev.length?D.chartCum35Prev[i]:'',
      i<D.chartRainCurr.length?D.chartRainCurr[i]:'',
      i<D.chartRainPrev.length?D.chartRainPrev[i]:'',
    ]);
  }
  const ws4 = XLSX.utils.aoa_to_sheet([dailyHeader, ...dailyRows]);
  XLSX.utils.book_append_sheet(wb, ws4, 'Daily Data');

  // Sheet 5: Heating Rate (weekly)
  const hrHeader = ['Week','Avg Max '+yr,'Avg Max '+ly,'5yr Normal'];
  const hrRows = [];
  const hrLen = Math.max((D.chartRateCurr||[]).length,(D.chartRatePrev||[]).length,(D.chartRateNorm||[]).length);
  for(let i=0;i<hrLen;i++){
    hrRows.push([i+1,
      D.chartRateCurr&&i<D.chartRateCurr.length?D.chartRateCurr[i]:'',
      D.chartRatePrev&&i<D.chartRatePrev.length?D.chartRatePrev[i]:'',
      D.chartRateNorm&&i<D.chartRateNorm.length?D.chartRateNorm[i]:'',
    ]);
  }
  const ws5 = XLSX.utils.aoa_to_sheet([hrHeader, ...hrRows]);
  XLSX.utils.book_append_sheet(wb, ws5, 'Heating Rate');

  XLSX.writeFile(wb, 'summer_strength_monitor_' + D.year + '.xlsx');
}

// ── INIT ──
drawChart('avg');renderTable();renderOnset();renderHeatmap();renderRegions();
window.addEventListener('resize',()=>drawChart(currentView));
</script>
</body>
</html>'''

# ═══════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ═══════════════════════════════════════════════════════════════
def generate_excel(stats, dates):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import datetime as dt_mod
    except ImportError:
        print("  openpyxl not installed — run: pip install openpyxl")
        return None

    yr, ly = dates["year"], dates["last_year"]
    wb = Workbook()

    # ── Shared styles ──
    HDR_FILL   = PatternFill("solid", fgColor="1C1917")
    HDR_FONT   = Font(name="Arial", bold=True, color="FAFAF9",  size=9)
    BODY_FONT  = Font(name="Arial", size=9)
    BOLD_FONT  = Font(name="Arial", bold=True, size=9)
    TITLE_FONT = Font(name="Arial", bold=True, color="FCD34D",  size=12)
    MUTED_FONT = Font(name="Arial", italic=True, color="A8A29E", size=8)

    S_FILL = PatternFill("solid", fgColor="14532D"); S_FONT = Font(name="Arial", bold=True, color="4ADE80", size=9)
    W_FILL = PatternFill("solid", fgColor="7F1D1D"); W_FONT = Font(name="Arial", bold=True, color="F87171", size=9)
    N_FILL = PatternFill("solid", fgColor="44403C"); N_FONT = Font(name="Arial", bold=True, color="A8A29E", size=9)

    def hdr(ws, row, cols, vals, fill=None, font=None):
        fill = fill or HDR_FILL; font = font or HDR_FONT
        for col, val in zip(cols, vals):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill; c.font = font
            c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

    def body_row(ws, row, col_start, vals):
        for i, val in enumerate(vals):
            c = ws.cell(row=row, column=col_start + i, value=val)
            c.font = BODY_FONT
            c.alignment = Alignment(horizontal="center", vertical="center")

    def set_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def signal_cell(ws, row, col, sig):
        c = ws.cell(row=row, column=col)
        if sig == "STRONGER":   c.fill = S_FILL; c.font = S_FONT
        elif sig == "WEAKER":   c.fill = W_FILL; c.font = W_FONT
        else:                   c.fill = N_FILL; c.font = N_FONT
        c.alignment = Alignment(horizontal="center")

    # ── Pan-India aggregates (reused across sheets) ──
    pac  = round(sa([s["avg_max_curr"]   for s in stats]), 1)
    mpac = round(sa([s["mar_avg_curr"]   for s in stats]), 1)
    mpap = round(sa([s["mar_avg_prev"]   for s in stats]), 1)
    mpan = round(sa([s["mar_avg_normal"] for s in stats if s["mar_avg_normal"]]), 1)
    mpd  = round(mpac - mpap, 1)
    mpdev = round(mpac - mpan, 1) if mpan else 0
    pssi = round(sa([s["ssi"] for s in stats]), 3)
    ns = sum(1 for s in stats if s["signal"] == "STRONGER")
    nw = sum(1 for s in stats if s["signal"] == "WEAKER")
    ni = len(stats) - ns - nw
    ssi_sig = "STRONGER" if pssi > 1.05 else ("WEAKER" if pssi < 0.95 else "IN-LINE")
    t35c = sum(s["hot35_curr"] for s in stats)
    t35p = sum(s["hot35_prev"] for s in stats)
    t40c = sum(s["hot40_curr"] for s in stats)
    onset30d = [s["onset30_delta"] for s in stats if s["onset30_delta"] is not None]
    onset35d = [s["onset35_delta"] for s in stats if s["onset35_delta"] is not None]
    avg30 = round(sa(onset30d), 0) if onset30d else None
    avg35 = round(sa(onset35d), 0) if onset35d else None
    e30 = sum(1 for d in onset30d if d < 0)
    e35 = sum(1 for d in onset35d if d < 0)

    sorted_stats = sorted(stats, key=lambda s: s["ssi"], reverse=True)

    # ════════════════════════════════════════════════════
    # Sheet 1 — Summary
    # ════════════════════════════════════════════════════
    ws1 = wb.active; ws1.title = "Summary"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 38
    ws1.column_dimensions["B"].width = 18
    ws1.column_dimensions["C"].width = 12
    ws1.column_dimensions["D"].width = 46

    t = ws1.cell(row=1, column=1, value=f"Summer Strength Monitor — {yr} Season")
    t.font = TITLE_FONT; t.fill = PatternFill("solid", fgColor="0C0A09")
    ws1.merge_cells("A1:D1"); t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    g = ws1.cell(row=2, column=1,
        value=f"Generated: {dates.get('end_current','')}  |  {len(stats)} cities  |  Season: Jan–Jun {yr}  |  vs {ly}")
    g.font = MUTED_FONT; ws1.merge_cells("A2:D2"); g.alignment = Alignment(horizontal="center")

    # — Pan-India metrics block —
    hdr(ws1, 4, [1, 2], ["Metric", "Value"])
    ws1.row_dimensions[4].height = 18
    metrics = [
        ("Pan-India SSI (Mar–Jun)",            f"{pssi:.3f}"),
        ("SSI Signal",                         ssi_sig),
        (f"Avg Max Temp Mar–Jun {yr}",         f"{mpac}°C"),
        (f"Avg Max Temp Mar–Jun {ly}",         f"{mpap}°C"),
        ("YoY Delta (Mar–Jun)",                f"{mpd:+.1f}°C"),
        ("vs 5yr Normal (Mar–Jun)",            f"{mpdev:+.1f}°C"),
        ("Cities STRONGER / WEAKER / IN-LINE", f"{ns} / {nw} / {ni}"),
        (f"Total City-Days ≥35°C ({yr})",      str(t35c)),
        (f"Total City-Days ≥35°C ({ly})",      str(t35p)),
        (f"Total City-Days ≥40°C ({yr})",      str(t40c)),
    ]
    for r_off, (k, v) in enumerate(metrics, 5):
        lc = ws1.cell(row=r_off, column=1, value=k); lc.font = BODY_FONT
        vc = ws1.cell(row=r_off, column=2, value=v); vc.font = BOLD_FONT
        vc.alignment = Alignment(horizontal="center")
        if k == "SSI Signal":
            signal_cell(ws1, r_off, 2, v)

    # — Onset summary block —
    onset_row = 5 + len(metrics) + 1
    hdr(ws1, onset_row, [1, 2], ["Onset Metric", "Value"])
    onset_data = [
        ("Avg 30°C Onset Delta (days)",   f"{avg30:+.0f}" if avg30 is not None else "—"),
        ("Avg 35°C Onset Delta (days)",   f"{avg35:+.0f}" if avg35 is not None else "—"),
        ("Cities with Earlier 30°C Onset", f"{e30} / {len(onset30d)}"),
        ("Cities with Earlier 35°C Onset", f"{e35} / {len(onset35d)}"),
    ]
    for r_off, (k, v) in enumerate(onset_data, onset_row + 1):
        ws1.cell(row=r_off, column=1, value=k).font = BODY_FONT
        vc = ws1.cell(row=r_off, column=2, value=v); vc.font = BOLD_FONT
        vc.alignment = Alignment(horizontal="center")


    # ════════════════════════════════════════════════════
    # Sheet 2 — City Data
    # ════════════════════════════════════════════════════
    ws2 = wb.create_sheet("City Data")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "C2"
    hdrs2 = [
        "City", "Region",
        f"Avg Max {yr}°C", f"Avg Max {ly}°C", "YoY Δ°C", "vs Normal°C", "5yr Normal°C",
        f"30°C Onset {yr}", f"30°C Onset {ly}", "30°C Δdays",
        f"35°C Onset {yr}", f"35°C Onset {ly}", "35°C Δdays",
        f"Hot≥35 {yr}", f"Hot≥35 {ly}", f"Days≥40 {yr}", "SSI", "Signal",
    ]
    hdr(ws2, 1, range(1, len(hdrs2) + 1), hdrs2)
    ws2.row_dimensions[1].height = 32

    for i, s in enumerate(sorted_stats, 2):
        d30 = s["onset30_delta"] if s["onset30_delta"] is not None else ""
        d35 = s["onset35_delta"] if s["onset35_delta"] is not None else ""
        vals = [
            s["city"], s["region"],
            s["mar_avg_curr"] or s["avg_max_curr"],
            s["mar_avg_prev"] or s["avg_max_prev"],
            s["mar_delta"] if s.get("mar_delta") is not None else s["delta"],
            s["mar_dev"] if s["mar_dev"] is not None else (s["dev_from_normal"] or ""),
            s["mar_avg_normal"] or s["avg_normal"] or "",
            s["onset30_curr"] or "", s["onset30_prev"] or "", d30,
            s["onset35_curr"] or "", s["onset35_prev"] or "", d35,
            s["hot35_curr"], s["hot35_prev"], s["hot40_curr"],
            s["ssi"], s["signal"],
        ]
        body_row(ws2, i, 1, vals)
        ws2.cell(row=i, column=1).alignment = Alignment(horizontal="left")
        ws2.cell(row=i, column=2).alignment = Alignment(horizontal="left")
        # SSI colour
        sc = ws2.cell(row=i, column=17)
        if s["ssi"] > 1.05:   sc.fill = S_FILL; sc.font = S_FONT
        elif s["ssi"] < 0.95: sc.fill = W_FILL; sc.font = W_FONT
        # Signal colour
        signal_cell(ws2, i, 18, s["signal"])

    set_widths(ws2, [16, 8, 10, 10, 8, 9, 10, 12, 12, 8, 12, 12, 8, 8, 8, 8, 7, 10])

    # ════════════════════════════════════════════════════
    # Sheet 3 — Monthly Heatmap
    # ════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Monthly Heatmap")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "C2"
    months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"]

    hdr(ws3, 1, [1], ["City"])
    hdr(ws3, 1, [2], ["Region"])
    hdr(ws3, 1, range(3, 9),   [f"{m} Avg°C" for m in months],
        fill=PatternFill("solid", fgColor="1C3A1C"),
        font=Font(name="Arial", bold=True, color="4ADE80", size=9))
    hdr(ws3, 1, range(9, 15),  [f"{m} Norm°C" for m in months],
        fill=PatternFill("solid", fgColor="1E293B"),
        font=Font(name="Arial", bold=True, color="60A5FA", size=9))
    hdr(ws3, 1, range(15, 21), [f"{m} Dev°C" for m in months],
        fill=PatternFill("solid", fgColor="2D1B00"),
        font=Font(name="Arial", bold=True, color="FCD34D", size=9))
    ws3.row_dimensions[1].height = 24

    DEV_HI_FILL  = PatternFill("solid", fgColor="7F1D1D"); DEV_HI_FONT  = Font(name="Arial", bold=True, color="F87171", size=9)
    DEV_MED_FILL = PatternFill("solid", fgColor="431407"); DEV_MED_FONT = Font(name="Arial", bold=True, color="FB923C", size=9)
    DEV_LO_FILL  = PatternFill("solid", fgColor="1E3A5F"); DEV_LO_FONT  = Font(name="Arial", bold=True, color="60A5FA", size=9)

    for i, s in enumerate(sorted_stats, 2):
        ws3.cell(row=i, column=1, value=s["city"]).font = BODY_FONT
        ws3.cell(row=i, column=2, value=s["region"]).font = Font(name="Arial", color="A8A29E", size=9)
        for j, m in enumerate(range(1, 7), 3):
            ws3.cell(row=i, column=j,    value=s["monthly_curr"].get(m, "")).font = BODY_FONT
            ws3.cell(row=i, column=j+6,  value=s["monthly_norm"].get(m, "")).font = Font(name="Arial", color="78716C", size=9)
            dev = s["monthly_dev"].get(m)
            dc = ws3.cell(row=i, column=j + 12, value=dev if dev is not None else "")
            if dev is not None:
                if dev >= 3:      dc.fill = DEV_HI_FILL;  dc.font = DEV_HI_FONT
                elif dev >= 1.5:  dc.fill = DEV_MED_FILL; dc.font = DEV_MED_FONT
                elif dev <= -1.5: dc.fill = DEV_LO_FILL;  dc.font = DEV_LO_FONT
                else:             dc.font = BODY_FONT
            for col in [j, j + 6, j + 12]:
                ws3.cell(row=i, column=col).alignment = Alignment(horizontal="center")

    set_widths(ws3, [16, 8] + [9] * 18)

    # ════════════════════════════════════════════════════
    # Sheet 4 — Daily Data
    # ════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Daily Data")
    ws4.sheet_view.showGridLines = False
    ws4.freeze_panes = "C2"
    hdrs4 = [
        "Day", "Date",
        f"Avg Max {yr}°C", f"Avg Max {ly}°C", "5yr Normal°C",
        f"Cities≥35 {yr}", f"Cities≥35 {ly}",
        f"Cum≥30 {yr}", f"Cum≥35 {yr}",
        f"Cum≥30 {ly}", f"Cum≥35 {ly}",
        f"Avg Rain {yr}mm", f"Avg Rain {ly}mm",
    ]
    hdr(ws4, 1, range(1, len(hdrs4) + 1), hdrs4)
    ws4.row_dimensions[1].height = 24

    ndc = max(len(s["daily_max_curr"])      for s in stats)
    ndp = max(len(s["daily_max_prev_full"]) for s in stats)
    ndn = max(len(s["daily_normals_full"])  for s in stats)
    max_len = max(ndc, ndp, ndn)

    # Build season date list
    season_dates = []
    d_iter = dt_mod.date(yr, 1, 1)
    end_date = dt_mod.date(yr, 6, 30)
    while d_iter <= end_date:
        season_dates.append(d_iter.isoformat())
        d_iter += dt_mod.timedelta(days=1)

    def davg(key, n):
        return [round(sa([s[key][i] for s in stats if i < len(s[key])]), 1)
                if any(i < len(s[key]) and s[key][i] is not None for s in stats) else None
                for i in range(n)]
    def dcnt(key, n, thr):
        return [sum(1 for s in stats if i < len(s[key]) and s[key][i] is not None and s[key][i] >= thr)
                for i in range(n)]
    def dcum(key, n):
        return [sum(s[key][i] for s in stats if i < len(s[key])) for i in range(n)]

    ca_c  = davg("daily_max_curr", ndc);       ca_p  = davg("daily_max_prev_full", ndp)
    cn    = davg("daily_normals_full", ndn)
    ch_c  = dcnt("daily_max_curr", ndc, 35);   ch_p  = dcnt("daily_max_prev_full", ndp, 35)
    cr_c  = davg("daily_rain_curr", ndc);       cr_p  = davg("daily_rain_prev_full", ndp)
    cc30c = dcum("cum30_curr", ndc);            cc35c = dcum("cum35_curr", ndc)
    cc30p = dcum("cum30_prev_full", ndp);       cc35p = dcum("cum35_prev_full", ndp)

    for i in range(max_len):
        row = [
            i + 1,
            season_dates[i] if i < len(season_dates) else "",
            ca_c[i]  if i < len(ca_c)  else "",
            ca_p[i]  if i < len(ca_p)  else "",
            cn[i]    if i < len(cn)    else "",
            ch_c[i]  if i < len(ch_c)  else "",
            ch_p[i]  if i < len(ch_p)  else "",
            cc30c[i] if i < len(cc30c) else "",
            cc35c[i] if i < len(cc35c) else "",
            cc30p[i] if i < len(cc30p) else "",
            cc35p[i] if i < len(cc35p) else "",
            cr_c[i]  if i < len(cr_c)  else "",
            cr_p[i]  if i < len(cr_p)  else "",
        ]
        body_row(ws4, i + 2, 1, row)

    set_widths(ws4, [5, 12, 11, 11, 11, 10, 10, 10, 10, 10, 10, 11, 11])

    # ════════════════════════════════════════════════════
    # Sheet 5 — Heating Rate (Weekly)
    # ════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Heating Rate")
    ws5.sheet_view.showGridLines = False
    ws5.freeze_panes = "B2"
    hdrs5 = ["Week", f"Avg Max {yr}°C", f"Avg Max {ly}°C", "5yr Normal°C"]
    hdr(ws5, 1, range(1, 5), hdrs5)

    nw_c = max(len(s["weekly_curr"]) for s in stats)
    nw_p = max(len(s["weekly_prev"]) for s in stats)
    nw_n = max(len(s["weekly_norm"]) for s in stats)

    def wavg(key, n):
        return [round(sa([s[key][i] for s in stats if i < len(s[key])]), 1)
                if any(i < len(s[key]) and s[key][i] is not None for s in stats) else None
                for i in range(n)]

    hw_c = wavg("weekly_curr", nw_c)
    hw_p = wavg("weekly_prev", nw_p)
    hw_n = wavg("weekly_norm", nw_n)

    for i in range(max(nw_c, nw_p, nw_n)):
        body_row(ws5, i + 2, 1, [
            i + 1,
            hw_c[i] if i < len(hw_c) else "",
            hw_p[i] if i < len(hw_p) else "",
            hw_n[i] if i < len(hw_n) else "",
        ])

    set_widths(ws5, [6, 13, 13, 13])

    xls_path = Path(OUTPUT_FILE).with_name(f"summer_strength_monitor_{yr}.xlsx")
    try:
        wb.save(str(xls_path))
    except PermissionError:
        # File is open in Excel — save with timestamp suffix instead
        ts = datetime.now().strftime("%H%M%S")
        xls_path = Path(OUTPUT_FILE).with_name(f"summer_strength_monitor_{yr}_{ts}.xlsx")
        wb.save(str(xls_path))
        print(f"  (previous file was open in Excel — saved as {xls_path.name})")
    return xls_path


# ═══════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════
def main():
    global VC_KEY
    print("="*60+"\n  Summer Strength Monitor v4 — Data Fetch\n"+"="*60)
    val = "--validate" in sys.argv or "-v" in sys.argv
    if val: print("  [VALIDATE MODE]\n")

    # Parse Visual Crossing API key for Mumbai
    # Priority: --vc-key arg > env var > hardcoded value at top of file
    for i, arg in enumerate(sys.argv):
        if arg == "--vc-key" and i+1 < len(sys.argv):
            VC_KEY = sys.argv[i+1]
    if not VC_KEY:
        VC_KEY = os.environ.get("VC_KEY")
    if VC_KEY:
        print(f"  Visual Crossing key: ...{VC_KEY[-6:]} (Mumbai will use METAR station data)")
    else:
        print("  ⚠ No Visual Crossing key — Mumbai will be SKIPPED")
        print("    Get free key: https://www.visualcrossing.com/sign-up")
        print("    Then: python summer_monitor.py --vc-key YOUR_KEY")
        print("    Or:   set VC_KEY=YOUR_KEY (Windows) / export VC_KEY=YOUR_KEY (Mac/Linux)")
    print()

    dates = get_date_ranges()
    print(f"  Year:      {dates['year']} vs {dates['last_year']}")
    print(f"  Current:   {dates['start_current']} → {dates['end_current']} ({dates['days_tracked']} days)")
    print(f"  Compare:   {dates['start_last']} → {dates['end_last_compare']}")
    print(f"  Prev Full: {dates['start_last']} → {dates['end_last_full']}")
    print(f"  5yr Base:  {dates['baseline_start']} → {dates['baseline_end']} ({dates['baseline_label']})\n")

    print("Fetching data (Open-Meteo + Visual Crossing for Mumbai)...")
    raw = fetch_all(dates)
    print("\nComputing analytics...")
    stats = [s for s in (compute_stats(raw[c]) for c in raw) if s]
    if not stats: print("  ERROR: No data."); sys.exit(1)
    print(f"  {len(stats)} cities OK")

    pssi = sa([s["ssi"] for s in stats])
    pdev = sa([s["dev_from_normal"] for s in stats if s["dev_from_normal"] is not None])
    ns=sum(1 for s in stats if s["signal"]=="STRONGER")
    nw=sum(1 for s in stats if s["signal"]=="WEAKER")
    o30=[s["onset30_delta"] for s in stats if s["onset30_delta"] is not None]
    o35=[s["onset35_delta"] for s in stats if s["onset35_delta"] is not None]
    print(f"\n  Pan-India SSI (Mar-Jun): {pssi:.3f}")
    print(f"  vs 5yr Normal:           {'+' if pdev>=0 else ''}{pdev:.1f}°C")
    print(f"  Signals: {ns} STRONGER / {nw} WEAKER / {len(stats)-ns-nw} IN-LINE")
    if o30: print(f"  Avg 30°C onset delta:    {sa(o30):+.0f} days ({sum(1 for d in o30 if d<0)}/{len(o30)} earlier)")
    if o35: print(f"  Avg 35°C onset delta:    {sa(o35):+.0f} days ({sum(1 for d in o35 if d<0)}/{len(o35)} earlier)")

    if val:
        print(f"\n{'='*70}\n  VALIDATION REPORT\n{'='*70}")
        print(f"\n  {'City':<14} {'Source':<6} {'Peak':<6} {'MarAvg':<8} {'MarDev':<7} {'Onset35':<12} {'Signal'}")
        for s in sorted(stats,key=lambda x:x["peak_curr"],reverse=True):
            d=raw.get(s["city"],{})
            src=d.get("source","om")
            md=f"{s['mar_dev']:+.1f}" if s.get("mar_dev") is not None else "—"
            print(f"  {s['city']:<14} {src:<6} {s['peak_curr']:<6} {s['mar_avg_curr']:<8} {md:<7} {str(s['onset35_curr'] or '—'):<12} {s['signal']}")

    print(f"\nGenerating {OUTPUT_FILE}...")
    html = generate_html(stats, dates)
    p = Path(OUTPUT_FILE); p.write_text(html, encoding="utf-8")
    print(f"  Saved: {p.resolve()}")

    print(f"\nGenerating Excel export...")
    xp = generate_excel(stats, dates)
    if xp:
        print(f"  Saved: {Path(xp).resolve()}")
        try:
            import os; os.startfile(str(Path(xp).resolve()))
        except Exception:
            pass

    try:
        import webbrowser; webbrowser.open(str(p.resolve())); print("  Browser opened!")
    except: pass
    print(f"\n{'='*60}\n  Done! Re-run anytime to refresh.\n{'='*60}")

if __name__ == "__main__": main()
